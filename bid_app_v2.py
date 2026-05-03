# ╔══════════════════════════════════════════════════════════════╗
# ║          입찰 투찰전략 분석 시스템  v1.0                    ║
# ║  - 3가지 분석값: ①패턴 ②유사표본 ③트렌드                  ║
# ║  - 배포자: 낙찰이력 관리 / 사용자: 전략 조회               ║
# ╚══════════════════════════════════════════════════════════════╝

import streamlit as st
import pandas as pd
import numpy as np
import xlrd
import io
import json
import os
from datetime import datetime

# ── 페이지 설정 ────────────────────────────────────────────────
st.set_page_config(
    page_title="투찰전략 분석 시스템",
    page_icon="📊",
    layout="wide"
)

# ── 스타일 ─────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1a2744, #243260);
        color: white; padding: 20px 30px; border-radius: 10px;
        margin-bottom: 20px;
    }
    .card {
        background: #f8fafc; border: 1px solid #e2e8f0;
        border-radius: 10px; padding: 15px; margin: 8px 0;
    }
    .val-box {
        border-radius: 8px; padding: 10px 15px;
        font-weight: bold; font-size: 1.1em; text-align: center;
    }
    .val-pattern  { background:#dbeafe; color:#1d4ed8; }
    .val-similar  { background:#dcfce7; color:#15803d; }
    .val-trend    { background:#fef9c3; color:#854d0e; }
    .val-recommend{ background:#f3e8ff; color:#7c3aed; }
    .badge-up   { background:#dcfce7; color:#15803d; border-radius:5px; padding:2px 8px; }
    .badge-down { background:#fee2e2; color:#991b1b; border-radius:5px; padding:2px 8px; }
    .badge-flat { background:#f1f5f9; color:#475569; border-radius:5px; padding:2px 8px; }
    .warning    { background:#fef9c3; border-left:4px solid #f59e0b; padding:10px; border-radius:5px; }
    .success    { background:#dcfce7; border-left:4px solid #10b981; padding:10px; border-radius:5px; }
</style>
""", unsafe_allow_html=True)

# ── 데이터 저장 경로 ───────────────────────────────────────────
DATA_DIR = "data"
HISTORY_FILE = os.path.join(DATA_DIR, "낙찰이력.pkl")
PATTERN_FILE = os.path.join(DATA_DIR, "패턴통계.json")
os.makedirs(DATA_DIR, exist_ok=True)

# ── 낙찰이력 로드 ──────────────────────────────────────────────
@st.cache_data
def load_history():
    if os.path.exists(HISTORY_FILE):
        return pd.read_pickle(HISTORY_FILE)
    return None

def save_history(df):
    df.to_pickle(HISTORY_FILE)
    st.cache_data.clear()

# ── 분석 함수 ──────────────────────────────────────────────────
def analyze_pattern(org, df_c):
    """① 발주처 패턴 분석"""
    sub = df_c[df_c['발주기관'] == org]['예가/기초(0%)'].values
    if len(sub) < 5:
        return None
    n = len(sub)
    mean = np.mean(sub); std = np.std(sub)
    r5  = np.mean(sub[-5:])
    r10 = np.mean(sub[-10:]) if n >= 10 else mean
    autocorr = float(np.corrcoef(sub[:-1], sub[1:])[0,1]) if n >= 3 else 0
    coef = float(np.polyfit(np.arange(min(20,n)), sub[-min(20,n):], 1)[0])
    trend   = '↑상승' if coef>0.02 else '↓하락' if coef<-0.02 else '→횡보'
    pattern = '연속성' if autocorr>0.2 else '반전' if autocorr<-0.2 else '무작위'
    if std > 0.60:          w5,w10,wm = 0.35,0.30,0.35
    elif abs(autocorr)>0.2: w5,w10,wm = 0.50,0.30,0.20
    else:                   w5,w10,wm = 0.40,0.35,0.25
    pred = w5*r5 + w10*r10 + wm*mean
    lv = float(sub[-1])
    if pattern=='연속성': adj = lv*abs(autocorr)*0.2
    elif pattern=='반전': adj = -lv*abs(autocorr)*0.3
    else: adj = 0.0
    return {
        'pred': round(pred+adj, 4), 'trend': trend, 'pattern': pattern,
        'autocorr': round(autocorr,4), 'last_val': round(lv,4),
        'r5': round(r5,4), 'r10': round(r10,4), 'mean': round(mean,4),
        'std': round(std,4), 'n': n,
        'recent10': [round(float(v),4) for v in sub[-10:]]
    }

def analyze_similar(name, base_원, df_c):
    """② 유사 용역 표본 분석"""
    keywords = []
    for kw in ['PD','VLF','감리','진단','설계','측정']:
        if kw in name:
            keywords.append(kw)
    if not keywords:
        keywords = ['감리']
    mask = pd.Series([False]*len(df_c), index=df_c.index)
    for kw in keywords:
        mask = mask | df_c['공고명'].str.contains(kw, na=False)
    similar = df_c[mask & (df_c['기초금액'] >= base_원*0.5) & (df_c['기초금액'] <= base_원*1.5)]
    if len(similar) < 3:
        similar = df_c[mask & (df_c['기초금액'] >= base_원*0.3) & (df_c['기초금액'] <= base_원*2.0)]
    if len(similar) < 3:
        return None
    vals = similar['예가/기초(0%)'].values
    n = len(vals)
    weights = np.linspace(0.5, 1.5, n)
    w_mean = float(np.average(vals, weights=weights))
    companies = similar['업체수'].mean() if '업체수' in similar.columns else None
    return {
        'pred': round(w_mean, 4), 'n': n,
        'mean': round(float(np.mean(vals)), 4),
        'std':  round(float(np.std(vals)), 4),
        'avg_companies': round(float(companies), 1) if companies else None,
        'keywords': keywords
    }

def analyze_trend(org, df_c):
    """③ 최근 트렌드 분석"""
    sub = df_c[df_c['발주기관'] == org]
    vals = sub['예가/기초(0%)'].values
    if len(vals) < 5:
        return None
    recent_n = max(5, len(vals)//4)
    recent  = vals[-recent_n:]
    older   = vals[:-recent_n]
    recent_mean = float(np.mean(recent))
    older_mean  = float(np.mean(older)) if len(older) > 0 else recent_mean
    drift = recent_mean - older_mean
    recent3 = vals[-3:] if len(vals) >= 3 else vals
    pred = recent_mean + drift*0.3
    companies = sub['업체수'].tail(recent_n).mean() if '업체수' in sub.columns else None
    return {
        'pred': round(pred, 4), 'recent_mean': round(recent_mean, 4),
        'drift': round(drift, 4), 'recent_n': recent_n,
        'recent3_mean': round(float(np.mean(recent3)), 4),
        'avg_companies': round(float(companies), 1) if companies else None
    }

def recommend_range(a1, a2, a3):
    """3가지 분석 기반 권장구간 계산"""
    vals = [v['pred'] for v in [a1, a2, a3] if v]
    if not vals:
        return None, None
    mean_v = np.mean(vals)
    std_v  = np.std(vals) if len(vals) > 1 else 0.1
    return round(mean_v - std_v*0.5, 4), round(mean_v + std_v*0.5, 4)

# ── XLS 파싱 ───────────────────────────────────────────────────
def parse_xls(file_bytes):
    wb = xlrd.open_workbook(file_contents=file_bytes, ignore_workbook_corruption=True)
    ws = wb.sheets()[0]
    headers = [ws.cell_value(1, c) for c in range(ws.ncols)]
    bids = []
    for r in range(2, ws.nrows):
        row = {headers[c]: ws.cell_value(r, c) for c in range(ws.ncols)}
        if not row.get('번호'):
            continue
        base = float(row.get('기초금액') or 0)
        bids.append({
            'no':       int(row['번호']),
            'name':     row.get('공고명',''),
            'bid_no':   row.get('공고번호',''),
            'base':     base,
            'base_억':  round(base/1e8, 4) if base else 0,
            'deadline': row.get('투찰마감',''),
            'org':      row.get('발주기관',''),
            'region':   row.get('지역',''),
        })
    return bids

# ── 엑셀 다운로드 생성 ─────────────────────────────────────────
def make_excel(results):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY='FF1a2744'; BLUE='FFdbeafe'; GREEN='FFdcfce7'; AMBER='FFfef9c3'
    RED_L='FFfee2e2'; PURP='FFf3e8ff'; GRAY='FFf8fafc'; WHITE='FFFFFFFF'
    thin=Side(style='thin',color='FFd1d5db')
    bdr=Border(left=thin,right=thin,top=thin,bottom=thin)

    def H(ws,r,c,v,bg=NAVY,fg='FFFFFFFF',sz=10,bold=True,wrap=False):
        cell=ws.cell(row=r,column=c,value=v)
        cell.font=Font(name='맑은 고딕',bold=bold,color=fg,size=sz)
        cell.fill=PatternFill('solid',start_color=bg)
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=wrap)
        cell.border=bdr; return cell

    def C(ws,r,c,v,bg=None,bold=False,right=False,sz=10,color='FF1e293b',center=False,wrap=False):
        cell=ws.cell(row=r,column=c,value=v)
        cell.font=Font(name='맑은 고딕',bold=bold,size=sz,color=color)
        ha='right' if right else ('center' if center else 'left')
        cell.alignment=Alignment(horizontal=ha,vertical='center',wrap_text=wrap)
        cell.border=bdr
        if bg: cell.fill=PatternFill('solid',start_color=bg)
        return cell

    wb = Workbook()
    ws = wb.active; ws.title='투찰전략'; ws.sheet_view.showGridLines=False

    today = datetime.now().strftime('%Y.%m.%d')
    ws.merge_cells('A1:L1')
    t=ws['A1']; t.value=f'투찰전략 분석표 — {today}  ★ 3가지 분석값 제공'
    t.font=Font(name='맑은 고딕',bold=True,size=13,color='FF1a2744')
    t.fill=PatternFill('solid',start_color='FFe0e7ff')
    t.alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height=30

    hdrs=['No','공고명','발주기관','기초금액(억)','마감',
          '①패턴(%)','②유사표본(%)','③트렌드(%)','권장하한(%)','권장상한(%)',
          '트렌드','패턴유형']
    wids=[5,42,22,11,13,11,11,11,11,11,9,10]
    for i,(h,w) in enumerate(zip(hdrs,wids),1):
        H(ws,2,i,h,wrap=True)
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[2].height=36

    for i,row in enumerate(results):
        r=i+3; bg=GRAY if r%2==0 else WHITE
        b=row['bid']; a1=row['a1']; a2=row['a2']; a3=row['a3']
        lo,hi=row['range_lo'],row['range_hi']

        C(ws,r,1,b['no'],bg=bg,bold=True,center=True)
        C(ws,r,2,b['name'][:55],bg=bg,sz=9,wrap=True)
        C(ws,r,3,b['org'],bg=bg,sz=9)
        if b['base']>0:
            cx=C(ws,r,4,b['base_억'],bg=bg,right=True); cx.number_format='#,##0.0000'
        else:
            C(ws,r,4,'미정',bg=bg,center=True,sz=9,color='FF9ca3af')
        C(ws,r,5,b['deadline'],bg=bg,sz=9,center=True)

        for ci,a,cbg in [(6,a1,BLUE),(7,a2,GREEN),(8,a3,AMBER)]:
            if a:
                cx2=C(ws,r,ci,a['pred'],bg=cbg,right=True,bold=True,
                      color='FF1d4ed8' if a['pred']>=0 else 'FF991b1b')
                cx2.number_format='+0.0000;-0.0000'
            else:
                C(ws,r,ci,'이력없음',bg=RED_L,center=True,sz=8,color='FF991b1b')

        for ci,val in [(9,lo),(10,hi)]:
            if val is not None:
                cx3=C(ws,r,ci,val,bg=PURP,right=True,bold=True,color='FF7c3aed')
                cx3.number_format='+0.0000;-0.0000'
            else:
                C(ws,r,ci,'-',bg=bg,center=True)

        if a1:
            tc='FF15803d' if '상승' in a1['trend'] else 'FF991b1b' if '하락' in a1['trend'] else 'FF64748b'
            C(ws,r,11,a1['trend'],bg=bg,center=True,color=tc,bold=True)
            pt_bg='FFccfbf1' if '연속' in a1['pattern'] else PURP if '반전' in a1['pattern'] else bg
            C(ws,r,12,a1['pattern'],bg=pt_bg,center=True,sz=9)
        ws.row_dimensions[r].height=34

    # 투찰금액 시트
    ws2=wb.create_sheet('투찰금액 환산'); ws2.sheet_view.showGridLines=False
    hdrs2=['No','공고명','기초금액(원)','①패턴 금액','②유사표본 금액','③트렌드 금액','권장하한 금액','권장상한 금액']
    wids2=[5,42,15,15,15,15,15,15]
    for i,(h,w) in enumerate(zip(hdrs2,wids2),1):
        H(ws2,1,i,h,wrap=True)
        ws2.column_dimensions[get_column_letter(i)].width=w
    ws2.row_dimensions[1].height=32

    for i,row in enumerate(results):
        r=i+2; bg=GRAY if r%2==0 else WHITE
        b=row['bid']; base=b['base']
        a1=row['a1']; a2=row['a2']; a3=row['a3']
        lo,hi=row['range_lo'],row['range_hi']

        C(ws2,r,1,b['no'],bg=bg,center=True,bold=True)
        C(ws2,r,2,b['name'][:55],bg=bg,sz=9,wrap=True)

        if base>0:
            cx=C(ws2,r,3,int(base),bg=bg,right=True); cx.number_format='#,##0'
            for ci,a,cbg in [(4,a1,BLUE),(5,a2,GREEN),(6,a3,AMBER)]:
                amt=int(base*(100+a['pred'])/100) if a else 0
                cbg2=cbg if a else RED_L
                cx2=C(ws2,r,ci,amt if a else '이력없음',bg=cbg2,right=True,bold=True,sz=10)
                if a: cx2.number_format='#,##0'
            for ci,val in [(7,lo),(8,hi)]:
                amt=int(base*(100+val)/100) if val is not None else None
                cx3=C(ws2,r,ci,amt,bg=PURP,right=True,bold=True); cx3.number_format='#,##0'
        else:
            for ci in range(3,9): C(ws2,r,ci,'기초금액 미정',bg=bg,center=True,sz=8,color='FF9ca3af')
        ws2.row_dimensions[r].height=28

    ws2.freeze_panes='A2'

    buf=io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf

# ════════════════════════════════════════════════════════════════
#  메인 UI
# ════════════════════════════════════════════════════════════════
st.markdown('<div class="main-header"><h2>📊 투찰전략 분석 시스템</h2><p style="margin:0;opacity:0.8">3가지 분석 기반 투찰전략 자동 산출 | v1.0</p></div>', unsafe_allow_html=True)

# ── 사이드바 ───────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ 시스템 설정")
    mode = st.radio("모드 선택", ["📊 투찰전략 분석", "🔧 배포자 관리"])
    st.divider()

    df_hist = load_history()
    if df_hist is not None:
        df_c = df_hist[df_hist['예가/기초(0%)'].notna() & (df_hist['예가/기초(0%)'].abs() < 10)].copy()
        st.success(f"✅ 낙찰이력 로드됨\n{len(df_c):,}건 | {df_c['발주기관'].nunique()}개 발주처")
    else:
        st.warning("⚠️ 낙찰이력 없음\n배포자 모드에서 데이터를 업로드하세요")
        df_c = None

# ══ 배포자 관리 모드 ══════════════════════════════════════════
if mode == "🔧 배포자 관리":
    st.header("🔧 배포자 관리 페이지")
    pwd = st.text_input("관리자 비밀번호", type="password")
    ADMIN_PWD = st.secrets.get("ADMIN_PWD", "admin1234")

    if pwd != ADMIN_PWD:
        st.info("비밀번호를 입력하면 관리 기능이 활성화됩니다.")
        st.stop()

    st.success("✅ 관리자 인증 완료")
    st.subheader("📂 낙찰이력 데이터 업로드")
    uploaded = st.file_uploader("낙찰이력 xlsx 파일을 업로드하세요", type=['xlsx','xls'])

    if uploaded:
        with st.spinner("데이터 처리 중..."):
            try:
                df_new = pd.read_excel(uploaded)
                required = ['발주기관','공고명','기초금액','예가/기초(0%)']
                missing = [c for c in required if c not in df_new.columns]
                if missing:
                    st.error(f"필수 컬럼 없음: {missing}")
                else:
                    save_history(df_new)
                    df_c2 = df_new[df_new['예가/기초(0%)'].notna() & (df_new['예가/기초(0%)'].abs() < 10)]
                    st.success(f"✅ 업로드 완료! {len(df_c2):,}건 / {df_c2['발주기관'].nunique()}개 발주처")

                    col1,col2,col3 = st.columns(3)
                    col1.metric("총 건수", f"{len(df_c2):,}건")
                    col2.metric("발주처 수", f"{df_c2['발주기관'].nunique()}개")
                    col3.metric("평균 사정율", f"{df_c2['예가/기초(0%)'].mean():+.4f}%")

                    st.subheader("📈 주요 발주처 현황")
                    top_orgs = df_c2['발주기관'].value_counts().head(15).reset_index()
                    top_orgs.columns=['발주기관','건수']
                    top_orgs['평균사정율'] = top_orgs['발주기관'].apply(
                        lambda x: round(df_c2[df_c2['발주기관']==x]['예가/기초(0%)'].mean(), 4))
                    st.dataframe(top_orgs, use_container_width=True, hide_index=True)
            except Exception as e:
                st.error(f"오류: {e}")

    if df_hist is not None:
        st.subheader("📊 현재 데이터 현황")
        df_c3 = df_hist[df_hist['예가/기초(0%)'].notna() & (df_hist['예가/기초(0%)'].abs() < 10)]
        st.info(f"현재 데이터: {len(df_c3):,}건 / {df_c3['발주기관'].nunique()}개 발주처")
        if st.button("🗑️ 데이터 초기화", type="secondary"):
            if os.path.exists(HISTORY_FILE):
                os.remove(HISTORY_FILE)
                st.cache_data.clear()
                st.success("초기화 완료")
                st.rerun()

# ══ 투찰전략 분석 모드 ════════════════════════════════════════
else:
    if df_c is None:
        st.error("낙찰이력 데이터가 없습니다. 배포자에게 문의하세요.")
        st.stop()

    st.header("📊 투찰전략 분석")

    # 파일 업로드
    col_up, col_info = st.columns([2,1])
    with col_up:
        xls_file = st.file_uploader(
            "입찰서류함 xls 파일을 업로드하세요",
            type=['xls','xlsx'],
            help="나라장터에서 다운받은 입찰서류함 파일"
        )
    with col_info:
        st.markdown("""
        <div class="card">
        <b>📌 분석 방법</b><br>
        ①패턴: 발주처 이력 패턴분석<br>
        ②유사표본: 유사용역 낙찰이력<br>
        ③트렌드: 최근 흐름 분석
        </div>
        """, unsafe_allow_html=True)

    if not xls_file:
        st.info("👆 입찰서류함 xls 파일을 업로드하면 자동으로 분석합니다.")
        st.stop()

    # 파일 파싱
    with st.spinner("파일 읽는 중..."):
        try:
            file_bytes = xls_file.read()
            bids = parse_xls(file_bytes)
            if not bids:
                st.error("입찰 건을 읽을 수 없습니다. 파일 형식을 확인해주세요.")
                st.stop()
        except Exception as e:
            st.error(f"파일 읽기 오류: {e}")
            st.stop()

    st.success(f"✅ {len(bids)}건 입찰 확인")
    st.divider()

    # 분석 실행
    results = []
    progress = st.progress(0, "분석 중...")

    for i, b in enumerate(bids):
        org = b['org']
        a1 = analyze_pattern(org, df_c)
        a2 = analyze_similar(b['name'], b['base'], df_c) if b['base'] > 0 else None
        a3 = analyze_trend(org, df_c)
        lo, hi = recommend_range(a1, a2, a3)
        results.append({'bid':b,'a1':a1,'a2':a2,'a3':a3,'range_lo':lo,'range_hi':hi})
        progress.progress((i+1)/len(bids), f"분석 중... {i+1}/{len(bids)}")

    progress.empty()

    # ── 결과 표시 ────────────────────────────────────────────
    st.subheader(f"📋 투찰전략 결과 — {datetime.now().strftime('%Y.%m.%d')}")

    # 요약 테이블
    summary_rows = []
    for row in results:
        b=row['bid']; a1=row['a1']; a2=row['a2']; a3=row['a3']
        lo,hi=row['range_lo'],row['range_hi']
        summary_rows.append({
            'No': b['no'],
            '공고명': b['name'][:40]+'...' if len(b['name'])>40 else b['name'],
            '발주기관': b['org'].replace('한국전력공사 ','한전 '),
            '기초금액(억)': f"{b['base_억']:.4f}" if b['base']>0 else '미정',
            '마감': b['deadline'],
            '①패턴(%)': f"{a1['pred']:+.4f}" if a1 else '이력없음',
            '②유사표본(%)': f"{a2['pred']:+.4f}" if a2 else '이력없음',
            '③트렌드(%)': f"{a3['pred']:+.4f}" if a3 else '이력없음',
            '권장하한(%)': f"{lo:+.4f}" if lo is not None else '-',
            '권장상한(%)': f"{hi:+.4f}" if hi is not None else '-',
            '트렌드': a1['trend'] if a1 else '-',
            '패턴': a1['pattern'] if a1 else '-',
        })

    df_summary = pd.DataFrame(summary_rows)
    st.dataframe(
        df_summary,
        use_container_width=True,
        hide_index=True,
        column_config={
            'No': st.column_config.NumberColumn(width=50),
            '공고명': st.column_config.TextColumn(width=280),
        }
    )

    st.divider()

    # 건별 상세 카드
    st.subheader("📌 건별 상세 분석")
    for row in results:
        b=row['bid']; a1=row['a1']; a2=row['a2']; a3=row['a3']
        lo,hi=row['range_lo'],row['range_hi']

        trend_badge = f'<span class="badge-{"up" if a1 and "상승" in a1["trend"] else "down" if a1 and "하락" in a1["trend"] else "flat"}">{a1["trend"] if a1 else "-"}</span>'

        with st.expander(f"No.{b['no']}  {b['name'][:55]}  |  {b['org'].replace('한국전력공사 ','한전 ')}  |  {b['base_억']:.4f}억  |  {b['deadline']}"):
            c1,c2,c3,c4 = st.columns(4)
            with c1:
                val_str = f"{a1['pred']:+.4f}%" if a1 else '이력없음'
                st.markdown(f'<div class="val-box val-pattern">①패턴<br>{val_str}</div>', unsafe_allow_html=True)
                if a1:
                    st.caption(f"n={a1['n']}건 | r5={a1['r5']:+.4f} | r10={a1['r10']:+.4f}")
                    st.caption(f"{a1['trend']} | {a1['pattern']}패턴 | 직전:{a1['last_val']:+.4f}%")
            with c2:
                val_str = f"{a2['pred']:+.4f}%" if a2 else '이력없음'
                st.markdown(f'<div class="val-box val-similar">②유사표본<br>{val_str}</div>', unsafe_allow_html=True)
                if a2:
                    st.caption(f"유사 {a2['n']}건 | 평균:{a2['mean']:+.4f}%")
                    companies = f"업체수 평균:{a2['avg_companies']}개" if a2['avg_companies'] else ""
                    st.caption(companies)
            with c3:
                val_str = f"{a3['pred']:+.4f}%" if a3 else '이력없음'
                st.markdown(f'<div class="val-box val-trend">③트렌드<br>{val_str}</div>', unsafe_allow_html=True)
                if a3:
                    st.caption(f"최근{a3['recent_n']}건 평균:{a3['recent_mean']:+.4f}%")
                    st.caption(f"drift:{a3['drift']:+.4f}% | 최근3건:{a3['recent3_mean']:+.4f}%")
            with c4:
                if lo is not None and hi is not None:
                    st.markdown(f'<div class="val-box val-recommend">💡권장구간<br>{lo:+.4f}%~{hi:+.4f}%</div>', unsafe_allow_html=True)
                    if b['base'] > 0:
                        st.caption(f"하한: {int(b['base']*(100+lo)/100):,}원")
                        st.caption(f"상한: {int(b['base']*(100+hi)/100):,}원")
                else:
                    st.markdown('<div class="val-box" style="background:#fee2e2;color:#991b1b;">⚠️데이터 부족</div>', unsafe_allow_html=True)

            # 최근 이력 시각화
            if a1 and a1['recent10']:
                st.caption("최근 10건 흐름:")
                vals = a1['recent10']
                colors = ['🟢' if v>=0 else '🔴' for v in vals]
                flow = ' '.join([f"{c}{v:+.3f}" for c,v in zip(colors, vals)])
                st.caption(flow + f" → 예측 🔵{a1['pred']:+.4f}")

    st.divider()

    # 다운로드
    st.subheader("💾 전략표 다운로드")
    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        excel_buf = make_excel(results)
        today_str = datetime.now().strftime('%Y%m%d')
        st.download_button(
            "📥 엑셀 다운로드 (투찰전략표)",
            data=excel_buf,
            file_name=f"투찰전략_{today_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )
    with col_dl2:
        # JSON 다운로드
        json_data = json.dumps([{
            'no':r['bid']['no'], 'name':r['bid']['name'],
            'org':r['bid']['org'], 'base':r['bid']['base'],
            'a1_pred':r['a1']['pred'] if r['a1'] else None,
            'a2_pred':r['a2']['pred'] if r['a2'] else None,
            'a3_pred':r['a3']['pred'] if r['a3'] else None,
            'range_lo':r['range_lo'], 'range_hi':r['range_hi'],
        } for r in results], ensure_ascii=False, indent=2)
        st.download_button(
            "📋 JSON 다운로드 (원시 데이터)",
            data=json_data.encode('utf-8'),
            file_name=f"투찰전략_{today_str}.json",
            mime="application/json",
            use_container_width=True
        )

    # 분석 통계
    with st.expander("📊 이번 분석 통계"):
        valid_a1 = [r for r in results if r['a1']]
        valid_a2 = [r for r in results if r['a2']]
        valid_a3 = [r for r in results if r['a3']]
        c1,c2,c3,c4 = st.columns(4)
        c1.metric("총 입찰건", f"{len(results)}건")
        c2.metric("①패턴 분석", f"{len(valid_a1)}/{len(results)}건")
        c3.metric("②유사표본", f"{len(valid_a2)}/{len(results)}건")
        c4.metric("③트렌드", f"{len(valid_a3)}/{len(results)}건")

