# ╔══════════════════════════════════════════════════════════════════╗
# ║  투찰전략 분석 시스템 v2.7                                      ║
# ║  개선: 백테스트 기반 추천모델 + 목표 적중률 추천구간             ║
# ║  - 비한전/조달청: 3포인트 미적용, 단일전략 표시                 ║
# ║  - ③트렌드 최소값 보정 (±0.02% 미만 시 보정)                  ║
# ║  - ②유사표본 없을 때 진단/감리 분야 전체평균으로 대체           ║
# ║  - 진단 분야 세분화 예측값 → ②유사표본에 통합                  ║
# ║  - 용역성격별 최적모델로 추천 중심값 산정                       ║
# ║  - 60/70/80% 목표 적중률별 추천 사정율 하한/상한                ║
# ╚══════════════════════════════════════════════════════════════════╝

import streamlit as st
import pandas as pd
import numpy as np
import xlrd, io, os, json
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from datetime import datetime

plt.rcParams.update({"font.family": "DejaVu Sans", "axes.unicode_minus": False})

st.set_page_config(page_title="투찰전략 분석 시스템", page_icon="📊", layout="wide")
st.markdown("""
<style>
.main-header{background:linear-gradient(135deg,#1a2744,#243260);color:white;
    padding:20px 30px;border-radius:10px;margin-bottom:20px}
.val-box{border-radius:8px;padding:10px 15px;font-weight:bold;
    font-size:1.1em;text-align:center;margin:4px 0}
.val-pattern{background:#dbeafe;color:#1d4ed8}
.val-similar{background:#dcfce7;color:#15803d}
.val-trend{background:#fef9c3;color:#854d0e}
.val-rec{background:#f3e8ff;color:#7c3aed}
.val-a{background:#fee2e2;color:#991b1b;border-radius:8px;padding:10px 15px;
    font-weight:bold;font-size:1.05em;text-align:center;margin:4px 0}
.val-b{background:#dbeafe;color:#1d4ed8;border-radius:8px;padding:10px 15px;
    font-weight:bold;font-size:1.05em;text-align:center;margin:4px 0}
.val-c{background:#dcfce7;color:#15803d;border-radius:8px;padding:10px 15px;
    font-weight:bold;font-size:1.05em;text-align:center;margin:4px 0}
.grade-a{background:#dcfce7;color:#15803d;padding:2px 8px;border-radius:4px;font-weight:bold}
.grade-b{background:#dbeafe;color:#1d4ed8;padding:2px 8px;border-radius:4px;font-weight:bold}
.grade-c{background:#fef9c3;color:#854d0e;padding:2px 8px;border-radius:4px;font-weight:bold}
.grade-d{background:#fee2e2;color:#991b1b;padding:2px 8px;border-radius:4px;font-weight:bold}
.three-pt-box{background:#f8fafc;border:2px solid #1a2744;border-radius:10px;
    padding:15px;margin:10px 0}
</style>""", unsafe_allow_html=True)

DATA_DIR     = "data"
HISTORY_FILE = os.path.join(DATA_DIR, "history.pkl")
PATTERN_FILE = os.path.join(DATA_DIR, "pattern_stats.json")
os.makedirs(DATA_DIR, exist_ok=True)

# ── 3포인트 전략 DB (분위수 기반, 30,614건 2026-05-13 업데이트) ─
# A = 하위25% 분위수 / B = 예측중심(차트예측) / C = 상위75% 분위수
# 실제 낙찰 분포는 ±1.5% 전 구간에 균등 분포 (복수예가 무작위 구조)
# ①②③ 예측은 중앙경향성 / A·C는 실제 분포 커버 역할
THREE_PT = {
    "한국전력공사 경기본부":         {"bias":"음수↓","detail":"음60%/중13%/양27%","pt_a":-0.43,"pt_c":+0.27,"cover":67,"cover_r":73,"lo95":-1.01,"hi95":+0.96,"note":"분위수기반"},
    "한국전력공사 부산울산본부":     {"bias":"음수↓","detail":"음55%/중23%/양23%","pt_a":-0.42,"pt_c":+0.29,"cover":71,"cover_r":87,"lo95":-1.07,"hi95":+0.91,"note":"최근커버 87%"},
    "한국전력공사 대전세종충남본부": {"bias":"음수↓","detail":"음39%/중29%/양32%","pt_a":-0.45,"pt_c":+0.25,"cover":73,"cover_r":77,"lo95":-1.05,"hi95":+0.80,"note":"대형발주처"},
    "한국전력공사 인천본부":         {"bias":"음수↓","detail":"음31%/중12%/양56%","pt_a":-0.40,"pt_c":+0.25,"cover":72,"cover_r":75,"lo95":-1.04,"hi95":+0.95,"note":"양수편향"},
    "한국전력공사 서울본부":         {"bias":"양수↑","detail":"음20%/중20%/양60%","pt_a":-0.44,"pt_c":+0.20,"cover":60,"cover_r":70,"lo95":-1.07,"hi95":+0.83,"note":"양수편향 강함"},
    "한국전력공사 경북본부":         {"bias":"균형", "detail":"음47%/중18%/양35%","pt_a":-0.38,"pt_c":+0.31,"cover":70,"cover_r":71,"lo95":-1.00,"hi95":+0.87,"note":"균형형"},
    "한국전력공사 경남본부":         {"bias":"음수↓","detail":"음50%/중30%/양20%","pt_a":-0.44,"pt_c":+0.26,"cover":71,"cover_r":100,"lo95":-1.12,"hi95":+0.81,"note":"최근커버 100%!"},
    "한국전력공사 광주전남본부":     {"bias":"음수↓","detail":"음52%/중10%/양38%","pt_a":-0.47,"pt_c":+0.24,"cover":68,"cover_r":67,"lo95":-1.08,"hi95":+0.86,"note":"음수편향"},
    "한국전력공사 대구본부":         {"bias":"음수↓","detail":"음52%/중12%/양36%","pt_a":-0.43,"pt_c":+0.24,"cover":71,"cover_r":72,"lo95":-1.08,"hi95":+0.84,"note":"음수편향"},
    "한국전력공사 강원본부":         {"bias":"음수↓","detail":"음53%/중13%/양33%","pt_a":-0.45,"pt_c":+0.24,"cover":65,"cover_r":93,"lo95":-1.05,"hi95":+0.83,"note":"최근커버 93%"},
    "한국전력공사 전북본부":         {"bias":"음수↓","detail":"음56%/중11%/양33%","pt_a":-0.48,"pt_c":+0.24,"cover":73,"cover_r":0, "lo95":-1.08,"hi95":+0.81,"note":"음수편향"},
    "한국전력공사 충북본부":         {"bias":"음수↓","detail":"음47%/중3%/양50%", "pt_a":-0.43,"pt_c":+0.28,"cover":72,"cover_r":0, "lo95":-1.01,"hi95":+0.88,"note":"균형형"},
    "한국전력공사 경기북부본부":     {"bias":"균형", "detail":"음53%/중11%/양37%","pt_a":-0.46,"pt_c":+0.30,"cover":70,"cover_r":58,"lo95":-1.07,"hi95":+0.87,"note":"균형형"},
    "한국전력공사 남서울본부":       {"bias":"균형", "detail":"음50%/중0%/양50%", "pt_a":-0.42,"pt_c":+0.29,"cover":72,"cover_r":0, "lo95":-1.03,"hi95":+0.92,"note":"균형형"},
    "한국전력공사 제주본부":         {"bias":"음수↓","detail":"음50%/중15%/양35%","pt_a":-0.45,"pt_c":+0.18,"cover":69,"cover_r":0, "lo95":-1.13,"hi95":+0.73,"note":"제주"},
}

# 3포인트 적용 대상 여부 판단
def is_three_pt_applicable(org, name):
    """한전 발주처 + 복수예가 방식인 경우만 3포인트 적용"""
    if not is_kepco(org):
        return False
    # 수의계약, 기초금액 매우 작은 건 제외
    if any(kw in name for kw in ['수의','소액수의','전자견적']):
        return False
    return True

# ── 기초금액 구간별 보정값 ─────────────────────────────────────
AMT_BRACKETS = {
    "~0.5억":  {"range":(0,   0.5),  "adj":+0.0013,"note":"소형"},
    "0.5~1억": {"range":(0.5, 1.0),  "adj":+0.0045,"note":"소형"},
    "1~2억":   {"range":(1.0, 2.0),  "adj":+0.0424,"note":"유리 ↑"},
    "2~5억":   {"range":(2.0, 5.0),  "adj":-0.0295,"note":"보수적 ↓"},
    "5~10억":  {"range":(5.0, 10.0), "adj":+0.0398,"note":"대형"},
    "10억+":   {"range":(10.0,9999), "adj":+0.0373,"note":"대형"},
}

def get_amt_info(base_억):
    for label, info in AMT_BRACKETS.items():
        lo, hi = info["range"]
        if lo <= base_억 < hi:
            return label, info["adj"], info["note"]
    return "미정", 0.0, ""

def get_three_pt(org, pred):
    """발주처별 3포인트 반환. 없으면 기본값"""
    if org in THREE_PT:
        d = THREE_PT[org]
        return {
            "pt_a": d["pt_a"],
            "pt_b": round(pred, 4),   # 차트예측값
            "pt_c": d["pt_c"],
            "bias": d["bias"],
            "detail": d["detail"],
            "cover": d["cover"],
            "cover_r": d["cover_r"],
            "note": d["note"],
            "found": True
        }
    return {
        "pt_a": -0.40, "pt_b": round(pred, 4), "pt_c": +0.20,
        "bias": "균형", "detail": "이력 부족",
        "cover": 60, "cover_r": 0, "note": "기본값 적용",
        "found": False
    }

# ── 낙찰이력 로드 ─────────────────────────────────────────────
@st.cache_data
def load_history():
    if os.path.exists(HISTORY_FILE):
        return pd.read_pickle(HISTORY_FILE)
    return None

@st.cache_data
def load_pattern_stats():
    if os.path.exists(PATTERN_FILE):
        with open(PATTERN_FILE, encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_history(df):
    os.makedirs(DATA_DIR, exist_ok=True)
    df.to_pickle(HISTORY_FILE)
    st.cache_data.clear()

def save_pattern_stats(stats):
    with open(PATTERN_FILE, "w", encoding="utf-8") as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)
    st.cache_data.clear()

# ── 분석 함수 ─────────────────────────────────────────────────
def analyze_pattern(org, df_c, pattern_stats):
    if org in pattern_stats:
        st_d = pattern_stats[org]
        return {
            "pred":         st_d.get("pred", 0),
            "conservative": st_d.get("conservative", 0),
            "aggressive":   st_d.get("aggressive", 0),
            "trend":        st_d.get("trend", "→횡보"),
            "pattern":      st_d.get("pattern", "무작위패턴").replace("패턴",""),
            "autocorr":     st_d.get("autocorr", 0),
            "last_val":     st_d.get("last_val", 0),
            "r5":           st_d.get("r5", 0),
            "r10":          st_d.get("r10", 0),
            "mean":         st_d.get("mean", 0),
            "std":          st_d.get("std", 0.5),
            "n":            st_d.get("n", 0),
            "w5":           st_d.get("w5", 0.25),
            "w10":          st_d.get("w10", 0.20),
            "wm":           st_d.get("wm", 0.55),
            "grade":        st_d.get("grade", "C"),
            "mae":          st_d.get("mae", 0.5),
            "drift6m":      st_d.get("drift6m", 0),
            "trend_boost":  st_d.get("trend_boost", 0),
            "drift_note":   st_d.get("drift_note", ""),
            "recent10":     st_d.get("recent10", []),
            "all_vals":     None,
            "source":       "패턴통계DB"
        }
    if df_c is None: return None
    sub = df_c[df_c["발주기관"]==org]["예가/기초(0%)"].values
    if len(sub) < 5: return None
    n=len(sub); mean=np.mean(sub); std=np.std(sub)
    r5=np.mean(sub[-5:]); r10=np.mean(sub[-10:]) if n>=10 else mean
    ac=float(np.corrcoef(sub[:-1],sub[1:])[0,1]) if n>=3 else 0
    coef=float(np.polyfit(np.arange(min(20,n)),sub[-min(20,n):],1)[0])
    trend   = "↑상승" if coef>0.02 else "↓하락" if coef<-0.02 else "→횡보"
    pattern = "연속성" if ac>0.2 else "반전" if ac<-0.2 else "무작위"
    w5,w10,wm = 0.25,0.20,0.55
    pred = w5*r5 + w10*r10 + wm*mean
    lv = float(sub[-1])
    adj = (lv*abs(ac)*0.2 if pattern=="연속성" else
           -lv*abs(ac)*0.3 if pattern=="반전" else 0.0)
    pred_final = round(pred+adj, 4)
    errs=[abs((0.25*np.mean(sub[:i][-5:])+0.20*(np.mean(sub[:i][-10:]) if i>=10 else np.mean(sub[:i]))+0.55*np.mean(sub[:i]))-sub[i])
          for i in range(min(10,n//2),n)]
    mae=np.mean(errs) if errs else 0.5
    grade="A" if mae<0.35 else "B" if mae<0.45 else "C" if mae<0.55 else "D"
    return {
        "pred":pred_final,"conservative":round(pred_final-std*0.4,4),
        "aggressive":round(pred_final+std*0.4,4),
        "trend":trend,"pattern":pattern,"autocorr":round(ac,4),
        "last_val":round(lv,4),"r5":round(r5,4),"r10":round(r10,4),
        "mean":round(mean,4),"std":round(std,4),"n":n,
        "w5":w5,"w10":w10,"wm":wm,"grade":grade,"mae":round(mae,4),
        "drift6m":0.0,"trend_boost":0.0,"drift_note":"",
        "recent10":[round(float(v),4) for v in sub[-10:]],
        "all_vals":sub.tolist(),"source":"직접계산"
    }

def analyze_similar(name, base_원, df_c):
    """② 유사표본 분석 — 이력 없으면 진단/감리 분야 전체평균으로 대체"""
    if df_c is None or base_원<=0: return None
    kws=[kw for kw in ["PD","VLF","감리","진단","설계","측정","광학","초음파","콘크리트"] if kw in name]
    if not kws: kws=["감리"]
    mask=pd.Series([False]*len(df_c),index=df_c.index)
    for kw in kws: mask=mask|df_c["공고명"].str.contains(kw,na=False)
    # ① 기초금액 ±50% 범위
    sim=df_c[mask&(df_c["기초금액"]>=base_원*0.5)&(df_c["기초금액"]<=base_원*1.5)]
    # ② ±100% 범위로 확대
    if len(sim)<3:
        sim=df_c[mask&(df_c["기초금액"]>=base_원*0.3)&(df_c["기초금액"]<=base_원*2.0)]
    # ③ 이력 부족 → 한전 포함 진단/감리 전체 평균으로 대체
    if len(sim)<3:
        if df_c is not None:
            # 진단 분야 → 한전 진단 전체평균
            if is_diag(name):
                kepco_mask = df_c['발주기관'].str.contains('한국전력공사',na=False)
                diag_mask  = df_c['공고명'].str.contains('|'.join(DIAG_KWS),na=False)
                sim = df_c[kepco_mask & diag_mask]
            # 감리 분야 → 한전 감리 전체평균
            elif is_supervision(name):
                kepco_mask = df_c['발주기관'].str.contains('한국전력공사',na=False)
                sup_mask   = df_c['공고명'].str.contains('감리',na=False)
                sim = df_c[kepco_mask & sup_mask]
        if len(sim)<3: return None
        # 대체 평균임을 표시
        vals=sim["예가/기초(0%)"].values; n=len(vals)
        weights=np.linspace(0.5,1.5,n)
        co=sim["업체수"].mean() if "업체수" in sim.columns else None
        return {
            "pred":round(float(np.average(vals,weights=weights)),4),
            "n":n,"mean":round(float(np.mean(vals)),4),
            "std":round(float(np.std(vals)),4),
            "avg_companies":round(float(co),1) if co and not np.isnan(float(co)) else None,
            "keywords":kws,"fallback":True,"fallback_note":"분야 전체평균 대체"
        }
    vals=sim["예가/기초(0%)"].values; n=len(vals)
    weights=np.linspace(0.5,1.5,n)
    co=sim["업체수"].mean() if "업체수" in sim.columns else None
    return {
        "pred":round(float(np.average(vals,weights=weights)),4),
        "n":n,"mean":round(float(np.mean(vals)),4),
        "std":round(float(np.std(vals)),4),
        "avg_companies":round(float(co),1) if co and not np.isnan(float(co)) else None,
        "keywords":kws,"fallback":False
    }

def analyze_trend(org, df_c):
    """③ 트렌드 분석 — 최소값 보정 (±0.02% 미만 시 평균으로 보정)"""
    if df_c is None: return None
    sub=df_c[df_c["발주기관"]==org]
    vals=sub["예가/기초(0%)"].values
    if len(vals)<5: return None
    rn=max(5,len(vals)//4); recent=vals[-rn:]; older=vals[:-rn]
    rm=float(np.mean(recent)); om=float(np.mean(older)) if len(older)>0 else rm
    drift=rm-om; r3=vals[-3:] if len(vals)>=3 else vals
    co=sub["업체수"].tail(rn).mean() if "업체수" in sub.columns else None
    raw_pred = rm+drift*0.3
    # ── 최소값 보정: 예측값이 ±0.02% 미만으로 0에 수렴하면 전체평균으로 대체
    if abs(raw_pred) < 0.02:
        raw_pred = float(np.mean(vals))
    return {
        "pred":round(raw_pred,4),"recent_mean":round(rm,4),
        "drift":round(drift,4),"recent_n":rn,
        "recent3_mean":round(float(np.mean(r3)),4),
        "avg_companies":round(float(co),1) if co and not np.isnan(float(co)) else None
    }

def recommend_range(a1,a2,a3):
    vals=[v["pred"] for v in [a1,a2,a3] if v]
    if not vals: return None,None
    mv=np.mean(vals); sv=np.std(vals) if len(vals)>1 else 0.1
    return round(mv-sv*0.5,4),round(mv+sv*0.5,4)

def convergence_score(a1,a2,a3):
    vals=[v["pred"] for v in [a1,a2,a3] if v]
    if len(vals)<2: return None,"데이터부족"
    sv=np.std(vals)
    if sv<0.05:   return sv,"★★★ 높음"
    elif sv<0.10: return sv,"★★☆ 보통"
    elif sv<0.20: return sv,"★☆☆ 낮음"
    else:         return sv,"⚠️ 분산큼"

# ── 개선 추천모델: 31,586건 이력 백테스트 기반 ─────────────────
SERVICE_LABELS = {
    "PD":"PD", "VLF":"VLF", "optical":"광학", "concrete":"콘크리트",
    "ultrasound":"초음파", "supervision":"감리", "design":"설계",
    "construction_management":"건설사업관리", "diagnosis_other":"진단기타",
    "other":"기타"
}

SERVICE_CHART_LABELS = {
    "PD":"PD", "VLF":"VLF", "optical":"Optical", "concrete":"Concrete",
    "ultrasound":"Ultrasound", "supervision":"Supervision", "design":"Design",
    "construction_management":"Construction Management",
    "diagnosis_other":"Diagnosis", "other":"Other"
}

MODEL_LABELS = {
    "svc_amount_recent20":"용역성격+금액구간 최근20건",
    "blend_weighted":"가중혼합",
    "service_nearest10_limited":"용역성격 유사패턴 최근10건",
    "org_service_recent20":"발주처+용역성격 최근20건",
    "blend_mean":"평균혼합",
    "org_service_nearest5_limited":"발주처+용역성격 유사패턴 최근5건",
    "org_recent20":"발주처 최근20건",
}

MODEL_RULES = {
    "PD": {"model":"svc_amount_recent20", "mae":0.4425, "p70":0.5960},
    "VLF": {"model":"blend_weighted", "mae":0.4038, "p70":0.5447},
    "concrete": {"model":"service_nearest10_limited", "mae":0.3614, "p70":0.4848},
    "construction_management": {"model":"org_service_recent20", "mae":0.5376, "p70":0.7028},
    "design": {"model":"blend_weighted", "mae":0.5174, "p70":0.6654},
    "diagnosis_other": {"model":"org_service_recent20", "mae":0.5035, "p70":0.6992},
    "optical": {"model":"blend_mean", "mae":0.4257, "p70":0.5555},
    "other": {"model":"org_service_nearest5_limited", "mae":0.3870, "p70":0.5000},
    "supervision": {"model":"org_service_recent20", "mae":0.4516, "p70":0.5842},
    "ultrasound": {"model":"org_recent20", "mae":0.5391, "p70":0.7099},
    "_default": {"model":"org_service_recent20", "mae":0.4787, "p70":0.6198},
}

TARGET_SCALE = {60:0.82, 70:1.00, 80:1.26}

def classify_service(name):
    s=str(name or "")
    su=s.upper()
    if "VLF" in su: return "VLF"
    if "PD" in su or "부분방전" in s: return "PD"
    if "광학" in s: return "optical"
    if "콘크리트" in s: return "concrete"
    if "초음파" in s: return "ultrasound"
    if "건설사업관리" in s or "감독권한대행" in s: return "construction_management"
    if "감리" in s: return "supervision"
    if "설계" in s: return "design"
    if "진단" in s or "점검" in s or "측정" in s: return "diagnosis_other"
    return "other"

def amount_bucket(base_원):
    try:
        eok=float(base_원 or 0)/1e8
    except Exception:
        eok=0
    if eok<=0: return "금액미상"
    if eok<0.3: return "0.3억미만"
    if eok<0.5: return "0.3~0.5억"
    if eok<1.0: return "0.5~1억"
    if eok<2.0: return "1~2억"
    if eok<5.0: return "2~5억"
    return "5억이상"

def company_bucket(v):
    try:
        x=float(v)
        if np.isnan(x): return None
        return int(x//10*10)
    except Exception:
        return None

def history_sorted(df):
    if df is None or len(df)==0: return pd.DataFrame()
    d=df.copy()
    for col in ["개찰일", "입찰일", "공고일", "마감", "투찰마감"]:
        if col in d.columns:
            dt=pd.to_datetime(d[col], errors="coerce")
            if dt.notna().sum()>0:
                d=d.assign(_dt=dt).sort_values("_dt")
                return d.drop(columns=["_dt"])
    return d.reset_index(drop=True)

def enrich_history(df):
    d=history_sorted(df)
    if len(d)==0: return d
    d=d.copy()
    d["_service"]=d["공고명"].apply(classify_service) if "공고명" in d.columns else "other"
    d["_amount_bucket"]=d["기초금액"].apply(amount_bucket) if "기초금액" in d.columns else "금액미상"
    d["_company_bucket"]=d["업체수"].apply(company_bucket) if "업체수" in d.columns else None
    return d

def val_mean(df, n=None):
    if df is None or len(df)==0 or "예가/기초(0%)" not in df.columns: return None
    vals=df["예가/기초(0%)"].dropna().astype(float)
    if len(vals)==0: return None
    if n: vals=vals.tail(n)
    return float(vals.mean()) if len(vals)>0 else None

def nearest_mean(df, current_hint, n=10, limit=180):
    if df is None or len(df)<3 or "예가/기초(0%)" not in df.columns: return None
    vals=df["예가/기초(0%)"].dropna().astype(float).tail(limit).to_numpy()
    if len(vals)<3: return None
    hint=float(current_hint if current_hint is not None else np.mean(vals[-min(10,len(vals)):]))
    idx=np.argsort(np.abs(vals-hint))[:min(n,len(vals))]
    return float(np.mean(vals[idx]))

def weighted_mean(items):
    ok=[(v,w) for v,w in items if v is not None]
    if not ok: return None
    sw=sum(w for _,w in ok)
    return float(sum(v*w for v,w in ok)/sw) if sw else None

def build_model_candidates(bid, df_e):
    if df_e is None or len(df_e)==0: return {}, {}
    org=bid.get("org","")
    svc=classify_service(bid.get("name",""))
    amt=amount_bucket(bid.get("base",0))
    base=df_e
    pools={
        "전체":base,
        "발주처":base[base["발주기관"].astype(str)==str(org)] if "발주기관" in base.columns else base.iloc[0:0],
        "용역성격":base[base["_service"]==svc] if "_service" in base.columns else base.iloc[0:0],
        "용역성격+금액":base[(base["_service"]==svc)&(base["_amount_bucket"]==amt)] if "_service" in base.columns else base.iloc[0:0],
        "발주처+용역성격":base[(base["발주기관"].astype(str)==str(org))&(base["_service"]==svc)] if "발주기관" in base.columns and "_service" in base.columns else base.iloc[0:0],
    }
    hint=val_mean(pools["발주처"], 10) or val_mean(pools["용역성격"], 10) or val_mean(base, 20)
    candidates={
        "all_recent20":val_mean(pools["전체"],20),
        "org_recent20":val_mean(pools["발주처"],20),
        "service_recent20":val_mean(pools["용역성격"],20),
        "svc_amount_recent20":val_mean(pools["용역성격+금액"],20),
        "org_service_recent20":val_mean(pools["발주처+용역성격"],20),
        "service_nearest10_limited":nearest_mean(pools["용역성격"], hint, n=10),
        "org_service_nearest5_limited":nearest_mean(pools["발주처+용역성격"], hint, n=5),
    }
    blend_base=[
        (candidates["org_recent20"],0.35),
        (candidates["org_service_recent20"],0.25),
        (candidates["svc_amount_recent20"] or candidates["service_recent20"],0.20),
        (candidates["service_recent20"],0.10),
        (candidates["all_recent20"],0.10),
    ]
    candidates["blend_weighted"]=weighted_mean(blend_base)
    mean_vals=[
        (candidates["org_recent20"],1),
        (candidates["org_service_recent20"],1),
        (candidates["svc_amount_recent20"] or candidates["service_recent20"],1),
        (candidates["all_recent20"],1),
    ]
    candidates["blend_mean"]=weighted_mean(mean_vals)
    return candidates, pools

def pick_candidate(model, candidates):
    fallback_order=[model, "org_service_recent20", "org_recent20", "svc_amount_recent20",
                    "service_recent20", "blend_weighted", "blend_mean", "all_recent20"]
    for key in fallback_order:
        val=candidates.get(key)
        if val is not None:
            return key, val
    return None, None

def analyze_improved_model(bid, df_c, target_hit=70):
    if df_c is None or len(df_c)==0: return None
    svc=classify_service(bid.get("name",""))
    rule=MODEL_RULES.get(svc, MODEL_RULES["_default"])
    df_e=df_c if "_service" in df_c.columns else enrich_history(df_c)
    candidates,pools=build_model_candidates(bid, df_e)
    model_used,pred=pick_candidate(rule["model"], candidates)
    if pred is None: return None
    scale=TARGET_SCALE.get(int(target_hit), 1.0)
    half=float(rule["p70"])*scale
    basis_pool=pools.get("발주처+용역성격")
    if basis_pool is None or len(basis_pool)<5:
        basis_pool=pools.get("용역성격")
    basis_n=len(basis_pool) if basis_pool is not None else len(df_e)
    model_label=MODEL_LABELS.get(model_used, model_used)
    svc_label=SERVICE_LABELS.get(svc, svc)
    return {
        "pred":round(float(pred),4),
        "lo":round(float(pred)-half,4),
        "hi":round(float(pred)+half,4),
        "service":svc,
        "service_label":svc_label,
        "model":model_used,
        "model_label":model_label,
        "target_hit":int(target_hit),
        "mae":round(float(rule["mae"]),4),
        "half_width":round(half,4),
        "basis_n":int(basis_n),
        "basis":f"{svc_label} 분류 후 백테스트 최적모델({model_label}) 적용",
        "candidates":{k:round(float(v),4) for k,v in candidates.items() if v is not None},
    }

def parse_xls(file_bytes, filename=""):
    """입찰서류함 파일 파싱 — xls/xlsx 모두 지원"""
    bids = []
    # ── xlsx 형식 ────────────────────────────────────────────
    is_xlsx = filename.lower().endswith(".xlsx") if filename else False
    if not is_xlsx:
        try:
            # xlsx 매직바이트 확인 (PK = zip 헤더)
            is_xlsx = file_bytes[:2] == b'PK'
        except Exception:
            is_xlsx = False

    if is_xlsx:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows_data = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows_data) < 3:
            return bids
        headers = [str(c) if c is not None else "" for c in rows_data[1]]
        for row_vals in rows_data[2:]:
            row = {headers[i]: row_vals[i] for i in range(len(headers))}
            if not row.get("번호"):
                continue
            try:
                base = float(row.get("기초금액") or 0)
            except (ValueError, TypeError):
                base = 0
            bids.append({
                "no":       int(float(row["번호"])),
                "name":     str(row.get("공고명") or ""),
                "bid_no":   str(row.get("공고번호") or ""),
                "base":     base,
                "base_억":  round(base/1e8, 4) if base else 0,
                "deadline": str(row.get("투찰마감") or ""),
                "org":      str(row.get("발주기관") or ""),
                "region":   str(row.get("지역") or ""),
            })
    else:
        # ── xls 형식 (나라장터 기본) ──────────────────────────
        wb = xlrd.open_workbook(file_contents=file_bytes, ignore_workbook_corruption=True)
        ws = wb.sheets()[0]
        headers = [ws.cell_value(1, c) for c in range(ws.ncols)]
        for r in range(2, ws.nrows):
            row = {headers[c]: ws.cell_value(r, c) for c in range(ws.ncols)}
            if not row.get("번호"):
                continue
            base = float(row.get("기초금액") or 0)
            bids.append({
                "no":       int(row["번호"]),
                "name":     row.get("공고명", ""),
                "bid_no":   row.get("공고번호", ""),
                "base":     base,
                "base_억":  round(base/1e8, 4) if base else 0,
                "deadline": row.get("투찰마감", ""),
                "org":      row.get("발주기관", ""),
                "region":   row.get("지역", ""),
            })
    return bids

# ── 영문 변환 ─────────────────────────────────────────────────
def tr_trend(t):
    return {"↑상승":"Up","↓하락":"Down","→횡보":"Flat"}.get(t,t)
def tr_pattern(p):
    return {"연속성":"Momentum","반전":"Reversal","무작위":"Random"}.get(p,p)
def tr_org(org):
    return (org.replace("한국전력공사 ","KEPCO ").replace("본부","")
               .replace("한국철도공사 회계통합센터","KORAIL")
               .replace("조달청","PPS").replace("국군재정관리단","MND"))

# ── 흐름 차트 ─────────────────────────────────────────────────
def make_flow_chart(a1,a2,a3,lo,hi,org_raw,three_pt=None):
    all_v=(a1.get("all_vals") or a1.get("recent10",[])) if a1 else []
    if not all_v: return None
    org_en=tr_org(org_raw)
    show_n=min(30,len(all_v)); recent=all_v[-show_n:]
    x=np.arange(1,show_n+1); next_x=show_n+1; mean_v=a1["mean"]

    fig=plt.figure(figsize=(13,5.5),facecolor="#f8fafc")
    gs=fig.add_gridspec(2,1,height_ratios=[2.6,1.0],hspace=0.04)
    ax=fig.add_subplot(gs[0]); ax_l=fig.add_subplot(gs[1])
    ax.set_facecolor("#ffffff"); ax_l.set_facecolor("#f8fafc"); ax_l.axis("off")

    ax.axhline(0,color="#94a3b8",lw=1.0,alpha=0.8,zorder=1)
    if lo is not None and hi is not None:
        ax.axhspan(lo,hi,alpha=0.12,color="#7c3aed",zorder=1)
        ax.axhline(lo,color="#7c3aed",lw=0.7,ls=":",alpha=0.5,zorder=1)
        ax.axhline(hi,color="#7c3aed",lw=0.7,ls=":",alpha=0.5,zorder=1)
    ax.axhline(mean_v,color="#f59e0b",lw=1.3,ls="--",alpha=0.85,zorder=2)
    ma5=[np.mean(recent[max(0,i-4):i+1]) for i in range(show_n)]
    ax.plot(x,ma5,color="#6366f1",lw=1.6,ls="--",alpha=0.75,zorder=3)
    bar_c=["#10b981" if v>=0 else "#ef4444" for v in recent]
    ax.bar(x,recent,color=bar_c,alpha=0.42,width=0.65,zorder=2)
    ax.plot(x,recent,color="#1a2744",lw=1.6,marker="o",ms=3.5,zorder=4)
    for i in range(max(0,show_n-10),show_n):
        v=recent[i]
        ax.annotate(f"{v:+.3f}",xy=(x[i],v),xytext=(0,7 if v>=0 else -11),
                    textcoords="offset points",ha="center",fontsize=6.5,
                    color="#059669" if v>=0 else "#dc2626",fontweight="bold")
    ax.axvline(show_n+0.5,color="#7c3aed",lw=1.4,ls=":",alpha=0.75)

    # 3포인트 표시
    if three_pt:
        pts=[
            ("A",three_pt["pt_a"],"#991b1b","v",10),
            ("B",three_pt["pt_b"],"#1d4ed8","D",11),
            ("C",three_pt["pt_c"],"#15803d","^",10),
        ]
        xoff=[-0.35,0.0,0.35]
        for idx,(lbl,pv,pc,mk,ms) in enumerate(pts):
            px=next_x+xoff[idx]
            ax.plot([px],[pv],color=pc,marker=mk,ms=ms,zorder=7,
                    markeredgecolor="white",markeredgewidth=1.2)
            ax.annotate(f"{lbl}:{pv:+.4f}%",xy=(px,pv),xytext=(22,0),
                        textcoords="offset points",ha="left",fontsize=8,
                        color=pc,fontweight="bold",
                        arrowprops=dict(arrowstyle="->",color=pc,lw=1.1))
        # 커버율 표시
        cover_txt=f"Cover:{three_pt['cover']}%"
        if three_pt['cover_r']>0: cover_txt+=f"(R:{three_pt['cover_r']}%)"
        ax.text(next_x+2.5,ax.get_ylim()[1]*0.9,cover_txt,
                fontsize=8,color="#7c3aed",fontweight="bold",ha="center",
                bbox=dict(boxstyle="round,pad=0.3",fc="#f3e8ff",ec="#7c3aed",alpha=0.85))
    else:
        preds=[]
        if a1: preds.append(("(1)Pattern",a1["pred"],"#1d4ed8","D",10))
        if a2: preds.append(("(2)Similar",a2["pred"],"#15803d","s",10))
        if a3: preds.append(("(3)Trend",  a3["pred"],"#92400e","^",10))
        xoff=[-0.32,0.0,0.32]
        for idx,(lbl,pv,pc,mk,ms) in enumerate(preds):
            px=next_x+xoff[idx]
            ax.plot([px],[pv],color=pc,marker=mk,ms=ms,zorder=7,
                    markeredgecolor="white",markeredgewidth=1.1)
            ax.annotate(f"{pv:+.4f}%",xy=(px,pv),xytext=(20,0),
                        textcoords="offset points",ha="left",fontsize=8,color=pc,
                        fontweight="bold",arrowprops=dict(arrowstyle="->",color=pc,lw=1.1))

    if lo is not None and hi is not None:
        bx=next_x+1.3
        ax.annotate("",xy=(bx,lo),xytext=(bx,hi),
                    arrowprops=dict(arrowstyle="<->",color="#7c3aed",lw=2.0))
        ax.text(bx+0.15,(lo+hi)/2,f"Rec.\n{lo:+.4f}\n~{hi:+.4f}",
                fontsize=7.5,color="#7c3aed",fontweight="bold",va="center",ha="left",
                bbox=dict(boxstyle="round,pad=0.3",fc="#f3e8ff",ec="#7c3aed",alpha=0.9))

    ax_r=ax.twinx(); ax_r.set_ylim(ax.get_ylim())
    tks=[mean_v]; tlbls=[f"Avg:{mean_v:+.3f}"]
    if lo is not None: tks+=[lo,hi]; tlbls+=[f"Lo:{lo:+.3f}",f"Hi:{hi:+.3f}"]
    ax_r.set_yticks(tks); ax_r.set_yticklabels(tlbls,fontsize=7,color="#64748b")
    ax.set_xlim(0.3,next_x+2.8)
    ax.set_xticks(list(x)+[next_x])
    ax.set_xticklabels([f"-{show_n-i}" for i in range(show_n)]+["Pred"],fontsize=7)
    ax.set_ylabel("Pred/Base(0%) %",fontsize=8)
    ax.tick_params(labelsize=7.5); ax.grid(axis="y",alpha=0.18,ls="--")
    grade=a1.get("grade","?"); mae=a1.get("mae",0)
    ax.set_title(
        f"{org_en}  |  Last {show_n} results  |  Trend:{tr_trend(a1['trend'])}"
        f"  |  Pattern:{tr_pattern(a1['pattern'])}  |  n={a1['n']}"
        f"  |  Grade:{grade}(MAE:{mae:.3f}%)",
        fontsize=9,fontweight="bold",color="#1a2744",pad=8)

    # 범례 패널
    ax_l.set_xlim(0,1); ax_l.set_ylim(0,1)
    ax_l.add_patch(mpatches.FancyBboxPatch(
        (0.005,0.03),0.990,0.94,boxstyle="round,pad=0.01",
        facecolor="#ffffff",edgecolor="#cbd5e1",linewidth=1.0,
        transform=ax_l.transAxes))
    col_defs=[(0.01,"Chart Legend","#1d4ed8"),(0.265,"Prediction","#15803d"),
              (0.515,"Pattern Detail","#7c3aed"),(0.765,"3-Point Strategy","#991b1b")]
    for cx,htxt,hcol in col_defs:
        ax_l.add_patch(mpatches.FancyBboxPatch(
            (cx+0.002,0.80),0.238,0.16,boxstyle="round,pad=0.005",
            facecolor=hcol,alpha=0.12,edgecolor="none",transform=ax_l.transAxes))
        ax_l.text(cx+0.012,0.885,htxt,fontsize=9,fontweight="bold",
                  color=hcol,va="center",transform=ax_l.transAxes)
    for lx in [0.255,0.505,0.755]:
        ax_l.plot([lx,lx],[0.04,0.97],color="#e2e8f0",lw=1.0,transform=ax_l.transAxes)

    a2n=a2["n"] if a2 else "-"
    grade_color={"A":"#15803d","B":"#1d4ed8","C":"#854d0e","D":"#991b1b"}.get(grade,"#475569")
    tp=three_pt or {}
    items=[
        (0.01,"line","#1a2744","","Actual bid result",  f"Last {show_n} results"),
        (0.01,"line","#6366f1","","Moving Avg MA(5)",    "5-case moving average"),
        (0.01,"line","#f59e0b","","Overall Average",     f"Avg:{mean_v:+.4f}%"),
        (0.01,"band","#7c3aed","","Recommended Zone",    f"{lo:+.4f}%~{hi:+.4f}%" if lo else "-"),
        (0.265,"mark","#1d4ed8","D","(1) Pattern",       f"w={a1['w5']}/{a1['w10']}/{a1['wm']} -> {a1['pred']:+.4f}%"),
        (0.265,"mark","#15803d","s","(2) Similar",       f"n={a2n} -> {a2['pred']:+.4f}%" if a2 else "No data"),
        (0.265,"mark","#92400e","^","(3) Trend",         f"{a3['pred']:+.4f}%" if a3 else "No data"),
        (0.515,"dot",grade_color,"","Accuracy Grade",    f"Grade:{grade} MAE:{mae:.3f}%"),
        (0.515,"dot","#7c3aed","","Last 5 avg (r5)",     f"{a1['r5']:+.4f}%"),
        (0.515,"dot","#7c3aed","","Last 10 avg (r10)",   f"{a1['r10']:+.4f}%"),
        (0.765,"mark","#991b1b","v","Co.A (음수/헷지)",  f"{tp.get('pt_a',0):+.2f}%"),
        (0.765,"mark","#1d4ed8","D","Co.B (차트예측)",   f"{tp.get('pt_b',0):+.4f}%"),
        (0.765,"mark","#15803d","^","Co.C (양수/헷지)",  f"{tp.get('pt_c',0):+.2f}%  Cover:{tp.get('cover',0)}%"),
    ]
    row_cnt={0.01:0,0.265:0,0.515:0,0.765:0}
    TOP_Y=0.72; ROW_GAP=0.225
    for (cx,itype,color,mk,label,desc) in items:
        ri=row_cnt[cx]; row_cnt[cx]+=1; y=TOP_Y-ri*ROW_GAP
        if itype=="line":
            ax_l.plot([cx+0.005,cx+0.038],[y+0.05,y+0.05],color=color,lw=2.4,
                      transform=ax_l.transAxes,clip_on=False)
        elif itype=="band":
            ax_l.add_patch(mpatches.FancyBboxPatch(
                (cx+0.005,y+0.02),0.033,0.065,boxstyle="round,pad=0.003",
                facecolor=color,alpha=0.28,edgecolor=color,linewidth=1.1,
                transform=ax_l.transAxes))
        elif itype in ("mark","dot"):
            ms_val=9 if itype=="mark" else 5.5
            ax_l.plot([cx+0.022],[y+0.05],marker=mk if itype=="mark" else "o",
                      color=color,ms=ms_val,transform=ax_l.transAxes,clip_on=False,
                      markeredgecolor="white" if itype=="mark" else color,
                      markeredgewidth=1.3 if itype=="mark" else 0,alpha=0.75 if itype=="dot" else 1)
        ax_l.text(cx+0.048,y+0.100,label,fontsize=8.5,fontweight="bold",
                  color="#1e293b",va="top",transform=ax_l.transAxes)
        ax_l.text(cx+0.048,y+0.005,desc,fontsize=8,color="#475569",
                  va="top",transform=ax_l.transAxes)
    ax_l.plot([0.01,0.99],[0.045,0.045],color="#e2e8f0",lw=0.8,transform=ax_l.transAxes)
    plt.subplots_adjust(left=0.055,right=0.92,top=0.95,bottom=0.02)
    buf=io.BytesIO()
    plt.savefig(buf,format="png",dpi=140,bbox_inches="tight",facecolor="#f8fafc")
    buf.seek(0); plt.close(); return buf

def simple_flow_data(df_c, bid, max_n=30):
    if df_c is None or len(df_c)==0: return None
    d=df_c if "_service" in df_c.columns else enrich_history(df_c)
    if len(d)==0 or "예가/기초(0%)" not in d.columns: return None
    org=str(bid.get("org",""))
    svc=classify_service(bid.get("name",""))
    org_df=d[d["발주기관"].astype(str)==org] if "발주기관" in d.columns else d.iloc[0:0]
    svc_df=org_df[org_df["_service"]==svc] if "_service" in org_df.columns else d.iloc[0:0]
    svc_scope="해당 발주처 내 해당분야"
    svc_chart_scope="Same agency service"
    if len(svc_df)<3 and "_service" in d.columns:
        svc_df=d[d["_service"]==svc]
        svc_scope="전체 발주처 해당분야"
        svc_chart_scope="All agencies service"
    org_vals=org_df["예가/기초(0%)"].dropna().astype(float).tail(max_n).to_list()
    svc_vals=svc_df["예가/기초(0%)"].dropna().astype(float).tail(max_n).to_list()
    if len(org_vals)==0 and len(svc_vals)==0: return None
    return {
        "org_label":"해당 발주처 전체",
        "svc_label":svc_scope,
        "service_label":SERVICE_LABELS.get(svc,svc),
        "org_chart_label":"Agency all",
        "svc_chart_label":svc_chart_scope,
        "service_chart_label":SERVICE_CHART_LABELS.get(svc,svc),
        "org_vals":org_vals,
        "svc_vals":svc_vals,
        "org_n":int(len(org_df)),
        "svc_n":int(len(svc_df)),
    }

def make_simple_flow_chart(df_c, bid, max_n=30):
    data=simple_flow_data(df_c,bid,max_n=max_n)
    if not data: return None, None
    fig,ax=plt.subplots(figsize=(12,3.8),facecolor="#ffffff")
    ax.set_facecolor("#ffffff")
    ax.axhline(0,color="#94a3b8",lw=1.0,alpha=0.9)

    def draw(vals,label,color,marker,label_pos="up"):
        if not vals: return
        x=np.arange(1,len(vals)+1)
        ax.plot(x,vals,color=color,lw=1.9,marker=marker,ms=4.5,label=f"{label} (n={len(vals)})")
        ax.scatter([x[-1]],[vals[-1]],s=50,color=color,zorder=5,edgecolor="white",linewidth=1.0)
        yoff=9 if label_pos=="up" else -13
        va="bottom" if label_pos=="up" else "top"
        for xi,yi in zip(x,vals):
            ax.annotate(f"{yi:+.4f}",xy=(xi,yi),xytext=(0,yoff),
                        textcoords="offset points",ha="center",va=va,
                        fontsize=6.5,color=color,fontweight="bold")

    draw(data["org_vals"],data["org_chart_label"],"#2563eb","o","up")
    draw(data["svc_vals"],data["svc_chart_label"],"#16a34a","s","down")
    ax.set_title(
        f"Adjustment Rate Trend - Agency vs {data['service_chart_label']}",
        fontsize=11,fontweight="bold",color="#1f2937",pad=10
    )
    ax.set_ylabel("Adjustment rate (%)",fontsize=9)
    ax.set_xlabel("Recent result order",fontsize=9)
    ax.grid(axis="y",alpha=0.22,ls="--")
    ax.tick_params(labelsize=8)
    ax.legend(loc="upper left",frameon=False,fontsize=9)
    ax.margins(x=0.03,y=0.18)
    plt.tight_layout()
    buf=io.BytesIO()
    plt.savefig(buf,format="png",dpi=140,bbox_inches="tight",facecolor="#ffffff")
    buf.seek(0); plt.close()
    return buf, data

# ── 한전 세분화 함수 ──────────────────────────────────────────
DIAG_KWS=['광학','초음파','VLF','PD','콘크리트']

def is_kepco(org): return '한국전력공사' in str(org)
def is_diag(name): return any(kw in str(name) for kw in DIAG_KWS)
def is_supervision(name): return '감리' in str(name)

def _sector_vals(df_c, org_filter, name_kws):
    if df_c is None: return np.array([])
    mask = df_c['발주기관'].str.contains(org_filter, na=False)
    kw_mask = pd.Series([False]*len(df_c), index=df_c.index)
    for kw in name_kws: kw_mask = kw_mask | df_c['공고명'].str.contains(kw, na=False)
    return df_c[mask & kw_mask]['예가/기초(0%)'].values

def _sector_stat(vals, scope):
    if len(vals)<3: return None
    n=len(vals); m=float(np.mean(vals)); s=float(np.std(vals))
    r5=float(np.mean(vals[-5:])) if n>=5 else m
    r10=float(np.mean(vals[-10:])) if n>=10 else m
    pred=0.25*r5+0.20*r10+0.55*m
    return {"pred":round(pred,4),"mean":round(m,4),"std":round(s,4),
            "r5":round(r5,4),"r10":round(r10,4),"n":n,
            "all_vals":vals.tolist(),"recent10":[round(float(v),4) for v in vals[-10:]],
            "conservative":round(pred-s*0.4,4),"aggressive":round(pred+s*0.4,4),
            "scope":scope}

def analyze_diag_all(df_c):
    v=_sector_vals(df_c,'한국전력공사','|'.join(DIAG_KWS).split('|'))
    return _sector_stat(v,'All KEPCO Diagnosis')

def analyze_diag_org(org,df_c):
    v=_sector_vals(df_c,org,DIAG_KWS)
    return _sector_stat(v,f"{org.replace('한국전력공사 ','').replace('본부','')} Diagnosis")

def analyze_sup_all(df_c):
    v=_sector_vals(df_c,'한국전력공사',['감리'])
    return _sector_stat(v,'All KEPCO Supervision')

def analyze_sup_org(org,df_c):
    v=_sector_vals(df_c,org,['감리'])
    return _sector_stat(v,f"{org.replace('한국전력공사 ','').replace('본부','')} Supervision")

def make_sector_chart(d_all,d_org,lo,hi,title):
    datasets=[]
    if d_all: datasets.append((d_all,d_all['scope'],'#1d4ed8'))
    if d_org: datasets.append((d_org,d_org['scope'],'#dc2626'))
    if not datasets: return None
    ncols=len(datasets)
    fig,axes=plt.subplots(1,ncols,figsize=(13,4.5),facecolor='#f8fafc',sharey=False)
    if ncols==1: axes=[axes]
    for ax,(data,scope,color) in zip(axes,datasets):
        ax.set_facecolor('#ffffff')
        vals=data.get('all_vals') or data.get('recent10',[])
        show_n=min(30,len(vals)); recent=vals[-show_n:]
        x=np.arange(1,show_n+1); mv=data['mean']; pv=data['pred']
        ax.axhline(0,color='#94a3b8',lw=1.0,alpha=0.7,zorder=1)
        ax.axhline(mv,color='#f59e0b',lw=1.3,ls='--',alpha=0.8,zorder=2)
        if lo and hi: ax.axhspan(lo,hi,alpha=0.10,color='#7c3aed',zorder=1)
        ma5=[np.mean(recent[max(0,i-4):i+1]) for i in range(show_n)]
        ax.plot(x,ma5,color='#6366f1',lw=1.5,ls='--',alpha=0.7,zorder=3)
        bar_c=[color if v>=0 else '#94a3b8' for v in recent]
        ax.bar(x,recent,color=bar_c,alpha=0.38,width=0.65,zorder=2)
        ax.plot(x,recent,color='#1a2744',lw=1.5,marker='o',ms=3.5,zorder=4)
        for i in range(max(0,show_n-10),show_n):
            v=recent[i]
            ax.annotate(f'{v:+.3f}',xy=(x[i],v),xytext=(0,7 if v>=0 else -11),
                        textcoords='offset points',ha='center',fontsize=6,
                        color='#059669' if v>=0 else '#dc2626',fontweight='bold')
        ax.axvline(show_n+0.5,color='#7c3aed',lw=1.2,ls=':',alpha=0.7)
        ax.plot([show_n+1],[pv],marker='D',color='#7c3aed',ms=9,zorder=7,
                markeredgecolor='white',markeredgewidth=1.2)
        ax.annotate(f'{pv:+.4f}%',xy=(show_n+1,pv),xytext=(16,0),
                    textcoords='offset points',ha='left',fontsize=8,color='#7c3aed',fontweight='bold',
                    arrowprops=dict(arrowstyle='->',color='#7c3aed',lw=1))
        ax.set_xlim(0.3,show_n+3.2)
        ax.set_xticks(list(x)+[show_n+1])
        ax.set_xticklabels([f'-{show_n-i}' for i in range(show_n)]+['Pred'],fontsize=7)
        ax.set_ylabel('Pred/Base(0%) %',fontsize=8)
        ax.tick_params(labelsize=7.5); ax.grid(axis='y',alpha=0.18,ls='--')
        ax.set_title(f'{scope}  (n={data["n"]})',fontsize=9.5,fontweight='bold',color='#1a2744',pad=6)
    fig.suptitle(title,fontsize=10,fontweight='bold',color='#1a2744',y=1.01)
    plt.tight_layout(pad=0.8)
    buf=io.BytesIO()
    plt.savefig(buf,format='png',dpi=130,bbox_inches='tight',facecolor='#f8fafc')
    buf.seek(0); plt.close(); return buf

# ── 엑셀 생성 ─────────────────────────────────────────────────
def make_excel(results):
    from openpyxl import Workbook
    from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
    from openpyxl.utils import get_column_letter
    NAVY="FF1a2744";BLUE="FFdbeafe";GREEN="FFdcfce7";AMBER="FFfef9c3"
    RED_L="FFfee2e2";PURP="FFf3e8ff";GRAY="FFf8fafc";RED2="FFfca5a5";GRN2="FFbbf7d0"
    thin=Side(style="thin",color="FFd1d5db"); bdr=Border(left=thin,right=thin,top=thin,bottom=thin)
    def H(ws,r,c,v,bg=NAVY,fg="FFFFFFFF",sz=10,bold=True,wrap=False):
        cell=ws.cell(row=r,column=c,value=v)
        cell.font=Font(name="맑은 고딕",bold=bold,color=fg,size=sz)
        cell.fill=PatternFill("solid",start_color=bg)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=wrap)
        cell.border=bdr; return cell
    def C(ws,r,c,v,bg=None,bold=False,right=False,sz=10,color="FF1e293b",center=False,wrap=False):
        cell=ws.cell(row=r,column=c,value=v)
        cell.font=Font(name="맑은 고딕",bold=bold,size=sz,color=color)
        ha="right" if right else ("center" if center else "left")
        cell.alignment=Alignment(horizontal=ha,vertical="center",wrap_text=wrap)
        cell.border=bdr
        if bg: cell.fill=PatternFill("solid",start_color=bg)
        return cell

    wb=Workbook(); ws=wb.active; ws.title="투찰전략"; ws.sheet_view.showGridLines=False
    today=datetime.now().strftime("%Y.%m.%d")
    ws.merge_cells("A1:W1"); t=ws["A1"]
    t.value=f"투찰전략 분석표 — {today}  ★ 백테스트 개선추천 + 3가지 분석값 + 3개업체 분산투찰 전략"
    t.font=Font(name="맑은 고딕",bold=True,size=13,color="FF1a2744")
    t.fill=PatternFill("solid",start_color="FFe0e7ff")
    t.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=30

    hdrs=["No","공고번호","공고명","발주기관","기초금액(억)","마감",
          "개선추천(%)","추천 사정율 하한(%)","추천 사정율 상한(%)","목표적중률",
          "적용모델","용역성격","모델평균오차","추천근거",
          "①패턴(%)","②유사표본(%)","③트렌드(%)","기존하한(%)","기존상한(%)",
          "업체A(%)","업체B(%)","업체C(%)","커버율"]
    wids=[5,18,44,20,10,12,12,14,14,10,22,12,12,34,10,10,10,10,10,10,10,10,8]
    for i,(h,w) in enumerate(zip(hdrs,wids),1):
        H(ws,2,i,h,wrap=True); ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[2].height=36

    for i,row in enumerate(results):
        r=i+3; bg=GRAY if r%2==0 else "FFFFFFFF"
        b=row["bid"]; a1=row["a1"]; a2=row["a2"]; a3=row["a3"]; im=row.get("improved")
        lo,hi=row["range_lo"],row["range_hi"]; tp=row.get("three_pt")
        C(ws,r,1,b["no"],bg=bg,bold=True,center=True)
        C(ws,r,2,b.get("bid_no",""),bg=bg,sz=9,center=True)
        C(ws,r,3,b["name"],bg=bg,sz=9,wrap=True)
        C(ws,r,4,b["org"],bg=bg,sz=9)
        if b["base"]>0:
            cx=C(ws,r,5,b["base_억"],bg=bg,right=True); cx.number_format="#,##0.0000"
        else: C(ws,r,5,"미정",bg=bg,center=True,sz=9)
        C(ws,r,6,b["deadline"],bg=bg,sz=9,center=True)
        if im:
            for ci,key in [(7,"pred"),(8,"lo"),(9,"hi")]:
                cx_im=C(ws,r,ci,im[key],bg=PURP,right=True,bold=True,color="FF7c3aed")
                cx_im.number_format="+0.0000;-0.0000"
            C(ws,r,10,f"{im['target_hit']}%",bg=PURP,center=True,bold=True,color="FF7c3aed")
            C(ws,r,11,im["model_label"],bg=PURP,sz=9,wrap=True)
            C(ws,r,12,im["service_label"],bg=PURP,center=True,sz=9)
            cx_mae=C(ws,r,13,im["mae"],bg=PURP,right=True); cx_mae.number_format="0.0000"
            C(ws,r,14,im["basis"],bg=PURP,sz=9,wrap=True)
        else:
            for ci in range(7,15): C(ws,r,ci,"-",bg=RED_L,center=True,sz=8)
        for ci,a,cbg in [(15,a1,BLUE),(16,a2,GREEN),(17,a3,AMBER)]:
            if a:
                cx2=C(ws,r,ci,a["pred"],bg=cbg,right=True,bold=True,
                      color="FF1d4ed8" if a["pred"]>=0 else "FF991b1b")
                cx2.number_format="+0.0000;-0.0000"
            else: C(ws,r,ci,"이력없음",bg=RED_L,center=True,sz=8)
        for ci,val in [(18,lo),(19,hi)]:
            if val is not None:
                cx3=C(ws,r,ci,val,bg=PURP,right=True,bold=True,color="FF7c3aed")
                cx3.number_format="+0.0000;-0.0000"
            else: C(ws,r,ci,"-",bg=bg,center=True)
        # 3포인트
        if tp:
            cx_a=C(ws,r,20,tp["pt_a"],bg=RED2,right=True,bold=True,color="FF991b1b")
            cx_a.number_format="+0.00;-0.00"
            cx_b=C(ws,r,21,tp["pt_b"],bg=BLUE,right=True,bold=True,color="FF1d4ed8")
            cx_b.number_format="+0.0000;-0.0000"
            cx_c=C(ws,r,22,tp["pt_c"],bg=GRN2,right=True,bold=True,color="FF15803d")
            cx_c.number_format="+0.00;-0.00"
            C(ws,r,23,f"{tp['cover']}%",bg=PURP,center=True,bold=True,color="FF7c3aed")
        else:
            for ci in [20,21,22,23]: C(ws,r,ci,"-",bg=bg,center=True)
        ws.row_dimensions[r].height=34

    # Sheet2: 3포인트 상세
    ws2=wb.create_sheet("3포인트 전략"); ws2.sheet_view.showGridLines=False
    ws2.merge_cells("A1:H1"); t2=ws2["A1"]
    t2.value="3개 업체 분산투찰 전략표 — 30,614건 기반"
    t2.font=Font(name="맑은 고딕",bold=True,size=12,color="FF1a2744")
    t2.fill=PatternFill("solid",start_color="FFe0e7ff")
    t2.alignment=Alignment(horizontal="center",vertical="center")
    ws2.row_dimensions[1].height=28
    hdrs2=["No","공고명","발주기관","기초금액","업체A포인트","업체B포인트(차트예측)","업체C포인트","커버율"]
    wids2=[5,40,20,12,14,18,14,10]
    for i,(h,w) in enumerate(zip(hdrs2,wids2),1):
        H(ws2,2,i,h,wrap=True); ws2.column_dimensions[get_column_letter(i)].width=w
    ws2.row_dimensions[2].height=30

    for i,row in enumerate(results):
        r=i+3; bg=GRAY if r%2==0 else "FFFFFFFF"
        b=row["bid"]; tp=row.get("three_pt")
        C(ws2,r,1,b["no"],bg=bg,center=True,bold=True)
        C(ws2,r,2,b["name"][:55],bg=bg,sz=9,wrap=True)
        C(ws2,r,3,b["org"],bg=bg,sz=9)
        if b["base"]>0:
            cx=C(ws2,r,4,b["base_억"],bg=bg,right=True); cx.number_format="#,##0.0000억"
        else: C(ws2,r,4,"미정",bg=bg,center=True)
        if tp:
            pa=tp["pt_a"]; pb=tp["pt_b"]; pc_v=tp["pt_c"]
            # 투찰금액
            if b["base"]>0:
                amt_a=int(b["base"]*(100+pa)/100)
                amt_b=int(b["base"]*(100+pb)/100)
                amt_c=int(b["base"]*(100+pc_v)/100)
                cxa=C(ws2,r,5,f"{pa:+.2f}% ({amt_a:,}원)",bg=RED2,bold=True,color="FF991b1b",center=True)
                cxb=C(ws2,r,6,f"{pb:+.4f}% ({amt_b:,}원)",bg=BLUE,bold=True,color="FF1d4ed8",center=True)
                cxc=C(ws2,r,7,f"{pc_v:+.2f}% ({amt_c:,}원)",bg=GRN2,bold=True,color="FF15803d",center=True)
            else:
                C(ws2,r,5,f"{pa:+.2f}%",bg=RED2,bold=True,color="FF991b1b",center=True)
                C(ws2,r,6,f"{pb:+.4f}%",bg=BLUE,bold=True,color="FF1d4ed8",center=True)
                C(ws2,r,7,f"{pc_v:+.2f}%",bg=GRN2,bold=True,color="FF15803d",center=True)
            cover_str=f"{tp['cover']}%"
            if tp['cover_r']>0: cover_str+=f"\n최근:{tp['cover_r']}%"
            C(ws2,r,8,cover_str,bg=PURP,bold=True,color="FF7c3aed",center=True,wrap=True)
        else:
            for ci in range(5,9): C(ws2,r,ci,"-",bg=bg,center=True)
        ws2.row_dimensions[r].height=36
    ws2.freeze_panes="A3"

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

# ════════════════════════════════════════════════════════════════
#  메인 UI
# ════════════════════════════════════════════════════════════════
st.markdown("""
<div class="main-header">
<h2>📊 투찰전략 분석 시스템 v2.7</h2>
<p style="margin:0;opacity:0.8">백테스트 기반 개선추천 + 기존 3가지 분석 + 3개 업체 분산투찰 전략</p>
</div>""", unsafe_allow_html=True)

with st.sidebar:
    st.header("⚙️ 시스템 설정")
    mode=st.radio("모드 선택",["📊 투찰전략 분석","🔧 배포자 관리"])
    target_hit=st.selectbox(
        "개선 추천구간 목표 적중률",
        [60,70,80],
        index=1,
        format_func=lambda x: f"{x}%형"
    )
    st.caption("목표 적중률이 높을수록 추천 사정율 하한/상한 폭이 넓어집니다.")
    st.divider()
    df_hist=load_history(); pattern_stats=load_pattern_stats()
    if df_hist is not None:
        df_c_s=df_hist[df_hist["예가/기초(0%)"].notna()&(df_hist["예가/기초(0%)"].abs()<10)]
        n_c=len(df_c_s); n_o=df_c_s["발주기관"].nunique()
        st.success(f"✅ 낙찰이력 {n_c:,}건\n{n_o}개 발주처")
    else:
        st.warning("⚠️ 낙찰이력 없음"); df_c_s=None; n_c=0; n_o=0
    if pattern_stats:
        st.success(f"✅ 패턴통계 {len(pattern_stats)}개 발주처")
    st.divider()
    st.caption(f"🕐 {datetime.now().strftime('%Y-%m-%d %H:%M')}")

# ══ 배포자 관리 ══════════════════════════════════════════════════
if mode=="🔧 배포자 관리":
    st.header("🔧 배포자 관리")
    pwd=st.text_input("관리자 비밀번호",type="password")
    ADMIN_PWD=st.secrets.get("ADMIN_PWD","admin1234")
    if pwd!=ADMIN_PWD: st.info("비밀번호를 입력하세요."); st.stop()
    st.success("✅ 관리자 인증")
    tab1,tab2=st.tabs(["📂 낙찰이력 업로드","📊 3포인트 전략 현황"])
    with tab1:
        uploaded=st.file_uploader("낙찰이력 xlsx 업로드",type=["xlsx","xls"])
        if uploaded:
            with st.spinner("처리 중..."):
                try:
                    content=uploaded.read()
                    df_new=pd.read_excel(io.BytesIO(content))
                    required=["발주기관","공고명","기초금액","예가/기초(0%)"]
                    missing=[c for c in required if c not in df_new.columns]
                    if missing: st.error(f"필수 컬럼 없음: {missing}")
                    else:
                        save_history(df_new)
                        df_v=df_new[df_new["예가/기초(0%)"].notna()&(df_new["예가/기초(0%)"].abs()<10)]
                        st.success("✅ 업로드 완료!")
                        c1,c2,c3=st.columns(3)
                        c1.metric("총 건수",f"{len(df_v):,}건")
                        c2.metric("발주처 수",f"{df_v['발주기관'].nunique()}개")
                        c3.metric("평균 사정율",f"{df_v['예가/기초(0%)'].mean():+.4f}%")
                except Exception as e: st.error(f"오류: {e}")
    with tab2:
        st.subheader("3포인트 전략 DB (30,614건 기반)")
        tp_rows=[]
        for org,d in THREE_PT.items():
            bias_icon="🔵" if "음수" in d['bias'] else "🔴" if "양수" in d['bias'] else "⚪"
            tp_rows.append({
                "발주처":org.replace("한국전력공사 ","한전 "),
                "편향":f"{bias_icon}{d['bias']}",
                "편향상세":d['detail'],
                "업체A":f"{d['pt_a']:+.2f}%",
                "업체C":f"{d['pt_c']:+.2f}%",
                "전체커버":f"{d['cover']}%",
                "최근커버":f"{d['cover_r']}%" if d['cover_r']>0 else "-",
                "비고":d['note']
            })
        st.dataframe(pd.DataFrame(tp_rows),use_container_width=True,hide_index=True)

# ══ 투찰전략 분석 ════════════════════════════════════════════════
else:
    df_hist=load_history()
    df_c = df_hist[df_hist["예가/기초(0%)"].notna()&(df_hist["예가/기초(0%)"].abs()<10)].copy() if df_hist is not None else None
    pattern_stats=load_pattern_stats()

    st.header("📊 투찰전략 분석")
    col_up,col_info=st.columns([2,1])
    with col_up:
        xls_file=st.file_uploader("입찰서류함 xls 파일 업로드",type=["xls","xlsx"])
    with col_info:
        nc=n_c if 'n_c' in dir() else (len(df_c) if df_c is not None else 0)
        no=n_o if 'n_o' in dir() else (df_c['발주기관'].nunique() if df_c is not None else 0)
        st.markdown(f"""
        <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:12px;font-size:0.9em">
        <b>📌 분석 방법</b><br>
        🔵 ①패턴: 발주처 이력 패턴분석<br>
        🟢 ②유사표본: 유사용역 낙찰이력<br>
        🟡 ③트렌드: 최근 흐름 분석<br>
        🟣 <b>개선추천: 백테스트 최적모델 추천값</b><br>
        🔴 <b>3포인트: A/B/C 분산투찰 전략</b><br><br>
        <b>데이터:</b> {nc:,}건 | {no}개 발주처
        </div>""",unsafe_allow_html=True)

    if not xls_file:
        st.info("👆 입찰서류함 xls 파일을 업로드하면 자동 분석합니다."); st.stop()

    with st.spinner("파일 읽는 중..."):
        try:
            raw_bytes = xls_file.read()
            # 낙찰이력 파일 오업로드 감지 (5MB 이상 + xlsx)
            if len(raw_bytes) > 3_000_000 and xls_file.name.lower().endswith(".xlsx"):
                st.error(
                    "⚠️ **잘못된 파일입니다.**\n\n"
                    "- 지금 업로드한 파일: **낙찰이력 데이터** (배포자 관리 전용)\n"
                    "- 여기서 필요한 파일: **나라장터 입찰서류함 xls**\n\n"
                    "👉 낙찰이력 업로드는 **사이드바 → 배포자 관리** 에서 진행하세요."
                )
                st.stop()
            bids = parse_xls(raw_bytes, xls_file.name)
            if not bids:
                st.error(
                    "입찰 건을 읽을 수 없습니다.\n\n"
                    "나라장터에서 다운받은 **입찰서류함 xls** 파일을 업로드해 주세요.\n"
                    "낙찰이력 파일(낙찰데이터.xlsx)은 **배포자 관리** 탭에서 업로드하세요."
                )
                st.stop()
        except Exception as e:
            st.error(
                f"파일 읽기 오류: {e}\n\n"
                "나라장터 입찰서류함 xls 파일인지 확인해 주세요."
            )
            st.stop()

    st.success(f"✅ {len(bids)}건 확인")
    results=[]
    df_model=enrich_history(df_c) if df_c is not None else None
    prog=st.progress(0,"분석 중...")
    for i,b in enumerate(bids):
        a1=analyze_pattern(b["org"],df_c,pattern_stats)
        a2=analyze_similar(b["name"],b["base"],df_c)
        a3=analyze_trend(b["org"],df_c)
        im=analyze_improved_model(b,df_model,target_hit)
        lo,hi=recommend_range(a1,a2,a3)
        conv_std,conv_lbl=convergence_score(a1,a2,a3)
        amt_lbl,amt_adj,amt_note=get_amt_info(b["base_억"])
        # 3포인트: 한전 + 수의계약 제외
        pred_val=a1["pred"] if a1 else 0.0
        tp=get_three_pt(b["org"],pred_val) if is_three_pt_applicable(b["org"],b["name"]) else None
        results.append({"bid":b,"a1":a1,"a2":a2,"a3":a3,
                        "improved":im,
                        "range_lo":lo,"range_hi":hi,
                        "conv_std":conv_std,"conv_lbl":conv_lbl,
                        "amt_lbl":amt_lbl,"amt_adj":amt_adj,"amt_note":amt_note,
                        "three_pt":tp})
        prog.progress((i+1)/len(bids))
    prog.empty()

    # ── 요약 테이블 ──────────────────────────────────────────
    st.subheader(f"📋 투찰전략 — {datetime.now().strftime('%Y.%m.%d')} ({len(bids)}건)")
    rows=[]
    for row in results:
        b=row["bid"]; a1=row["a1"]; a2=row["a2"]; a3=row["a3"]
        lo,hi=row["range_lo"],row["range_hi"]; tp=row["three_pt"]; im=row.get("improved")
        grade=a1.get("grade","?") if a1 else "?"
        ge={"A":"🟢","B":"🔵","C":"🟡","D":"🔴"}.get(grade,"⚪")
        tp_str=f"A:{tp['pt_a']:+.2f} B:{tp['pt_b']:+.4f} C:{tp['pt_c']:+.2f} ({tp['cover']}%)" if tp else "-"
        lo95_str=f"{tp['lo95']:+.2f}~{tp['hi95']:+.2f}%" if (tp and tp.get('lo95')) else "-"
        rows.append({"No":b["no"],
            "공고명":b["name"][:33]+"…" if len(b["name"])>33 else b["name"],
            "발주기관":b["org"].replace("한국전력공사 ","한전 "),
            "기초(억)":f"{b['base_억']:.4f}" if b["base"]>0 else "미정",
            "마감":b["deadline"],
            "①패턴":f"{a1['pred']:+.4f}%" if a1 else "없음",
            "②유사표본":f"{a2['pred']:+.4f}%" if a2 else "없음",
            "③트렌드":f"{a3['pred']:+.4f}%" if a3 else "없음",
            "개선추천":f"{im['pred']:+.4f}%" if im else "없음",
            "추천 사정율 하한":f"{im['lo']:+.4f}%" if im else "-",
            "추천 사정율 상한":f"{im['hi']:+.4f}%" if im else "-",
            "적용모델":im["model_label"] if im else "-",
            "기존하한":f"{lo:+.4f}%" if lo is not None else "-",
            "기존상한":f"{hi:+.4f}%" if hi is not None else "-",
            "분석값 일치도":row["conv_lbl"],"등급":f"{ge}{grade}",
            "3포인트":tp_str,
            "실분포95%":lo95_str})
    st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True,
                 column_config={"No":st.column_config.NumberColumn(width=50),
                                 "공고명":st.column_config.TextColumn(width=200),
                                "적용모델":st.column_config.TextColumn(width=180),
                                 "3포인트":st.column_config.TextColumn(width=220)})
    st.divider()

    # ── 건별 상세 ────────────────────────────────────────────
    st.subheader("📌 건별 상세 + 사정율 흐름 차트")
    for row in results:
        b=row["bid"]; a1=row["a1"]; a2=row["a2"]; a3=row["a3"]
        lo,hi=row["range_lo"],row["range_hi"]; tp=row["three_pt"]; im=row.get("improved")
        rec_lo,rec_hi=(im["lo"],im["hi"]) if im else (lo,hi)
        grade=a1.get("grade","?") if a1 else "?"
        ge={"A":"🟢","B":"🔵","C":"🟡","D":"🔴"}.get(grade,"⚪")
        label=(f"No.{b['no']}  {b['name'][:48]}  |  "
               f"{b['org'].replace('한국전력공사 ','한전 ')}  |  "
               f"{b['base_억']:.4f}억  |  {b['deadline']}  {ge}{grade}")
        with st.expander(label):
            # 신뢰도 + 금액구간 배너
            col_g,col_a=st.columns([1,2])
            with col_g:
                gc={"A":"grade-a","B":"grade-b","C":"grade-c","D":"grade-d"}.get(grade,"grade-c")
                mae_v=a1.get("mae",0) if a1 else 0
                st.markdown(f'<span class="{gc}">신뢰도 {grade} (MAE:{mae_v:.3f}%)</span>',unsafe_allow_html=True)
            with col_a:
                amt_note=row["amt_note"]; amt_adj=row["amt_adj"]; amt_lbl=row["amt_lbl"]
                if "유리" in amt_note: st.success(f"💰 {amt_lbl} — 유리한 구간 ({amt_adj:+.4f}%)")
                elif "보수" in amt_note: st.warning(f"⚠️ {amt_lbl} — 보수적 접근 권장 ({amt_adj:+.4f}%)")
                else: st.info(f"📐 {amt_lbl} ({amt_adj:+.4f}%)")

            c1,c2,c3,c4,c5=st.columns(5)
            with c1:
                v=f"{a1['pred']:+.4f}%" if a1 else "이력없음"
                st.markdown(f'<div class="val-box val-pattern">①패턴<br>{v}</div>',unsafe_allow_html=True)
                if a1:
                    st.caption(f"n={a1['n']}건 | {a1['trend']} | {a1['pattern']}패턴")
                    st.caption(f"r5={a1['r5']:+.4f} / r10={a1['r10']:+.4f} / 직전:{a1['last_val']:+.4f}%")
                    if a1.get("std",0) > 0.3:
                        st.caption(f"⚠️ std={a1['std']:.3f}% — 실제분포 넓음")
            with c2:
                v=f"{a2['pred']:+.4f}%" if a2 else "이력없음"
                fb = a2.get("fallback",False) if a2 else False
                box_style = "val-similar" if not fb else "val-trend"
                st.markdown(f'<div class="val-box {box_style}">②유사표본{"(대체)" if fb else ""}<br>{v}</div>',unsafe_allow_html=True)
                if a2:
                    if fb:
                        st.caption(f"⚠️ {a2.get('fallback_note','분야평균 대체')}")
                    st.caption(f"유사 {a2['n']}건 | 평균:{a2['mean']:+.4f}%")
                    if a2.get("avg_companies"): st.caption(f"업체수 참고:{a2['avg_companies']}개")
            with c3:
                v=f"{a3['pred']:+.4f}%" if a3 else "이력없음"
                st.markdown(f'<div class="val-box val-trend">③트렌드<br>{v}</div>',unsafe_allow_html=True)
                if a3:
                    st.caption(f"최근{a3['recent_n']}건평균:{a3['recent_mean']:+.4f}%")
                    st.caption(f"drift:{a3['drift']:+.4f}% | 최근3건:{a3['recent3_mean']:+.4f}%")
            with c4:
                if im:
                    st.markdown(f'<div class="val-box val-rec">개선추천<br>{im["pred"]:+.4f}%</div>',unsafe_allow_html=True)
                    st.caption(f"{im['model_label']} | {im['target_hit']}%형")
                    st.caption(f"추천 사정율 하한/상한: {im['lo']:+.4f}% ~ {im['hi']:+.4f}%")
                    st.caption(f"평균오차 {im['mae']:.4f}%p | 근거 {im['basis_n']}건")
                else:
                    st.markdown('<div class="val-box" style="background:#fee2e2;color:#991b1b">개선추천<br>데이터부족</div>',unsafe_allow_html=True)
            with c5:
                if lo is not None:
                    st.markdown(f'<div class="val-box val-rec">기존 권장<br>{lo:+.4f}%~{hi:+.4f}%</div>',unsafe_allow_html=True)
                    if b["base"]>0:
                        st.caption(f"하한: {int(b['base']*(100+lo)/100):,}원")
                        st.caption(f"상한: {int(b['base']*(100+hi)/100):,}원")
                    st.caption(f"분석값 일치도: {row['conv_lbl']}")
                else:
                    st.markdown('<div class="val-box" style="background:#fee2e2;color:#991b1b">⚠️ 데이터부족</div>',unsafe_allow_html=True)

            # ── 3포인트 분산투찰 카드 (한전만) ───────────────
            if tp:
                st.markdown("---")
                bias_icon="🔵" if "음수" in tp["bias"] else "🔴" if "양수" in tp["bias"] else "⚪"
                cover_r_str=f" | 최근커버: **{tp['cover_r']}%**" if tp['cover_r']>0 else ""
                lo95=tp.get("lo95",0); hi95=tp.get("hi95",0)
                st.markdown(
                    f"**🏢 3개 업체 분산투찰 전략** &nbsp;"
                    f"{bias_icon} 편향: **{tp['bias']}** ({tp['detail']}) &nbsp;|&nbsp; "
                    f"커버율: **{tp['cover']}%**{cover_r_str}",
                    unsafe_allow_html=False
                )
                # ── 복수예가 분포 안내 ──────────────────────────
                if lo95 and hi95:
                    st.info(
                        f"📊 **복수예가 실제 분포 안내** — "
                        f"실제 낙찰은 **{lo95:+.2f}% ~ {hi95:+.2f}%** 전 구간에 균등 분포합니다 (95% 범위). "
                        f"①②③ 예측값은 중앙경향성(평균 근처)이며, "
                        f"업체A·C는 실제 분포의 **하위25%({tp['pt_a']:+.2f}%)·상위75%({tp['pt_c']:+.2f}%)** 구간을 커버합니다."
                    )
                ca,cb,cc=st.columns(3)
                with ca:
                    amt_a=f"\n{int(b['base']*(100+tp['pt_a'])/100):,}원" if b['base']>0 else ""
                    st.markdown(f'<div class="val-a">🏢 업체A<br>{tp["pt_a"]:+.2f}%{amt_a}</div>',unsafe_allow_html=True)
                    st.caption(f"▶ 하위25% 분위수 | {'음수주력' if '음수' in tp['bias'] else '음수헷지'}")
                    if b['base']>0: st.caption(f"※ 수익성 검토 필요")
                with cb:
                    amt_b=f"\n{int(b['base']*(100+tp['pt_b'])/100):,}원" if b['base']>0 else ""
                    st.markdown(f'<div class="val-b">🏢 업체B (차트예측)<br>{tp["pt_b"]:+.4f}%{amt_b}</div>',unsafe_allow_html=True)
                    st.caption("▶ 중앙경향성 예측값 (핵심 포인트)")
                with cc:
                    amt_c=f"\n{int(b['base']*(100+tp['pt_c'])/100):,}원" if b['base']>0 else ""
                    st.markdown(f'<div class="val-c">🏢 업체C<br>{tp["pt_c"]:+.2f}%{amt_c}</div>',unsafe_allow_html=True)
                    st.caption(f"▶ 상위75% 분위수 | {'양수주력' if '양수' in tp['bias'] else '양수헷지'}")

            # ── 사정율 흐름 차트: 발주처 전체 vs 해당분야 ───────
            st.markdown("---")
            st.markdown("**사정율 흐름 차트**")
            with st.spinner("단순 비교차트 생성 중..."):
                chart_buf,chart_data=make_simple_flow_chart(df_model,b,max_n=30)
            if chart_buf:
                m1,m2=st.columns(2)
                with m1:
                    vals=chart_data["org_vals"]
                    st.metric(
                        "해당 발주처 전체",
                        f"{np.mean(vals):+.4f}%" if vals else "-",
                        f"표본 {chart_data['org_n']}건"
                    )
                with m2:
                    vals=chart_data["svc_vals"]
                    st.metric(
                        f"{chart_data['service_label']} 분야",
                        f"{np.mean(vals):+.4f}%" if vals else "-",
                        f"{chart_data['svc_label']} {chart_data['svc_n']}건"
                    )
                st.image(chart_buf,use_container_width=True)
            else:
                st.caption("⚠️ 차트에 표시할 이력 데이터가 부족합니다.")

    st.divider()
    st.subheader("💾 전략표 다운로드")
    excel_buf=make_excel(results)
    today_str=datetime.now().strftime("%Y%m%d")
    st.download_button("📥 엑셀 다운로드 (v2.1 — 3포인트 전략 포함)",
        data=excel_buf,
        file_name=f"투찰전략_{today_str}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",use_container_width=True)
